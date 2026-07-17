import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
import csv
import os

wb = openpyxl.load_workbook(
    r"c:\Users\warren\Documents\GitHub\Warrenhughes1974\docs\Segment References.xlsx",
    data_only=False,
)


def classify_fill(cell):
    fill = cell.fill
    if not fill or fill.fill_type is None:
        return "none"
    fg = fill.fgColor
    if fg is None:
        return "none"
    if fg.type == "rgb" and fg.rgb:
        rgb = str(fg.rgb).upper()
        rgb6 = rgb[2:] if len(rgb) == 8 else rgb
    else:
        theme = fg.theme if fg.type == "theme" else None
        tint = getattr(fg, "tint", 0) or 0
        if theme == 5:
            return "peach"
        if theme == 8:
            return "purple"
        if theme == 0:
            return "row_header_gray"
        return f"theme_{theme}_{tint:.4f}"

    mapping = {
        "00B050": "dark_green",
        "92D050": "light_green",
        "FF0000": "red",
        "FFFF00": "yellow",
        "FFC000": "peach",
        "FCE4D6": "peach",
        "F8CBAD": "peach",
        "E2EFDA": "light_green",
        "C6EFCE": "light_green",
        "7030A0": "purple",
        "B4A7D6": "purple",
        "D9A7E7": "purple",
        "CCCCFF": "purple",
        "FF99CC": "peach",
    }
    if rgb6 in mapping:
        return mapping[rgb6]
    return f"rgb_{rgb6}"


def cell_val(c):
    v = c.value
    return "" if v is None else str(v).strip()


all_rows = []
color_counts = Counter()
for name in wb.sheetnames:
    ws = wb[name]
    headers = {}
    for c in range(1, ws.max_column + 1):
        headers[c] = cell_val(ws.cell(1, c)) or get_column_letter(c)
    for r in range(2, ws.max_row + 1):
        policy = cell_val(ws.cell(r, 1))
        if not policy:
            continue
        for c in range(2, ws.max_column + 1):
            cell = ws.cell(r, c)
            v = cell_val(cell)
            if not v:
                continue
            cat = classify_fill(cell)
            color_counts[cat] += 1
            all_rows.append(
                {
                    "sheet": name,
                    "row": r,
                    "policy_form": policy,
                    "rate_type_col": headers[c],
                    "col": get_column_letter(c),
                    "segment": v,
                    "assessment": cat,
                }
            )

print("COLOR COUNTS:", dict(color_counts))
print("TOTAL valued cells:", len(all_rows))

by_assess = defaultdict(list)
for row in all_rows:
    by_assess[row["assessment"]].append(row)

order = [
    "dark_green",
    "light_green",
    "purple",
    "red",
    "peach",
    "yellow",
    "none",
] + [
    k
    for k in by_assess
    if k
    not in {
        "dark_green",
        "light_green",
        "purple",
        "red",
        "peach",
        "yellow",
        "none",
        "row_header_gray",
    }
]

for assess in order:
    items = by_assess.get(assess, [])
    if not items:
        continue
    print(f"\n### {assess} ({len(items)}) ###")
    for it in items:
        print(
            f"  [{it['sheet']}] {it['policy_form']} | {it['rate_type_col']} | {it['segment']}"
        )

out = r"c:\Users\warren\Documents\GitHub\Warrenhughes1974\Issue_Log_Items\Issue_42\evidence_segment_references_parsed.csv"
os.makedirs(os.path.dirname(out), exist_ok=True)
with open(out, "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=list(all_rows[0].keys()))
    w.writeheader()
    w.writerows(all_rows)
print("\nWrote", out)
