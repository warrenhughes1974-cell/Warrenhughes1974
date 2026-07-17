"""Inventory PDAGE keys missing from Rate_Table and crosswalk resolve."""
from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(r"c:\Users\warren\Documents\GitHub\Warrenhughes1974")
SRC = ROOT / "QLA_Migration" / "Source"
PDAGE = SRC / "PDAGE_AgeDuration_Rates_Extract_20260713.csv"
RT = SRC / "Rate_Table_Extract_Txt.txt"
XW = ROOT / "plan_analysis" / "source_data" / "crosswalk" / "Policy Form Crosswalk 5.22.26.xlsx"
OUT = ROOT / "Issue_Log_Items" / "Issue_42" / "evidence_pdage_missfill_inventory.csv"


def norm(row):
    return {(k or "").strip(): (v or "").strip() for k, v in row.items()}


def main():
    rt_ct: dict[tuple[str, str], int] = defaultdict(int)
    with open(RT, newline="", encoding="utf-8", errors="replace") as f:
        for raw in csv.DictReader(f):
            r = norm(raw)
            cov, typ = r.get("COVERAGE_ID", ""), r.get("TYPE_CODE", "")
            if cov and set(cov) != {"-"}:
                rt_ct[(cov, typ)] += 1

    pd_ct: dict[tuple[str, str], int] = defaultdict(int)
    with open(PDAGE, newline="", encoding="utf-8", errors="replace") as f:
        for raw in csv.DictReader(f):
            r = norm(raw)
            cov, typ = r.get("COVERAGE_ID", ""), r.get("TYPE_CODE", "")
            if cov and set(cov) != {"-"}:
                pd_ct[(cov, typ)] += 1

    miss = [(c, t, n) for (c, t), n in sorted(pd_ct.items()) if rt_ct.get((c, t), 0) == 0]

    wb = openpyxl.load_workbook(XW, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cov2plan = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] and row[2]:
            cov2plan[str(row[0]).strip()] = str(row[2]).strip()

    print(f"PDAGE unique (cov,type)={len(pd_ct)}")
    print(f"Rate_Table unique (cov,type)={len(rt_ct)}")
    print(f"Miss-fill keys={len(miss)} rows={sum(x[2] for x in miss):,}")

    focus = ["L01 10Y", "L10 LP9595", "L17", "960 LP85-8", "L01 10Y LT", "L10 LP95"]
    print("\nFocus plan resolve:")
    for c in focus:
        print(f"  {c!r} -> {cov2plan.get(c, 'UNRESOLVED')}")

    by_type = defaultdict(int)
    by_type_rows = defaultdict(int)
    rows_out = []
    for cov, typ, n in miss:
        by_type[typ] += 1
        by_type_rows[typ] += n
        plan = cov2plan.get(cov, "")
        rows_out.append(
            {
                "coverage_id": cov,
                "type_code": typ,
                "pdage_rows": n,
                "plan": plan or "UNRESOLVED",
                "resolvable": "Y" if plan else "N",
            }
        )

    print("\nMiss-fill by TYPE:")
    for t in sorted(by_type):
        print(f"  {t}: {by_type[t]} keys, {by_type_rows[t]:,} rows")

    resolved = [r for r in rows_out if r["resolvable"] == "Y"]
    unresolved = [r for r in rows_out if r["resolvable"] == "N"]
    print(
        f"\nResolved keys={len(resolved)} rows={sum(int(r['pdage_rows']) for r in resolved):,}"
    )
    print(
        f"UNRESOLVED keys={len(unresolved)} rows={sum(int(r['pdage_rows']) for r in unresolved):,}"
    )

    print("\nIssue42 / Eric focus miss keys:")
    for r in rows_out:
        c = r["coverage_id"]
        if c in (
            "L01 10Y",
            "L10 LP9595",
            "L17",
            "960 LP85-8",
            "0824 P DTH",
            "L10 GPO OL",
        ) or "LP85" in c or c.startswith("L17"):
            print(
                f"  {c} {r['type_code']} rows={r['pdage_rows']} plan={r['plan']}"
            )

    print("\nTop 25 unresolved by rows:")
    for r in sorted(unresolved, key=lambda x: -int(x["pdage_rows"]))[:25]:
        print(f"  {r['coverage_id']} {r['type_code']} {r['pdage_rows']}")

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["coverage_id", "type_code", "pdage_rows", "plan", "resolvable"],
        )
        w.writeheader()
        w.writerows(rows_out)
    print(f"\nWrote {OUT}")


if __name__ == "__main__":
    main()
