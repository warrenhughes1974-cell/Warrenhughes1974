"""
QLAdmin V5 rate factor loader — stage + transform + pivot LifePRO rate extracts into the
QLAdmin wide factor grid. READ-ONLY with respect to inputs; produces in-memory structures.

Pipeline:
  1. load_plan_crosswalk()  COVERAGE_ID -> authoritative PLAN (Policy Form Crosswalk).
  2. transform_source()     classify rows (in-scope / excluded / rejected), apply confirmed
                            SEX/BAND/UWCLASS crosswalks + duration->CNTL paging.
  3. build_factor_grid()    pivot transformed cells into factor rows keyed by the confirmed
                            segmentation tuple, each carrying the xx0..xx9 columns.

No DBFs are written here. Segmentation defaults (EFFDATE/ISSCNTRY/ISSUEST) are supplied via
config (externalized) and NEVER invented inline.
"""
import csv
import collections

from qla_core import rate_dbf_schema as S

SOURCE_COLUMNS = ("COVERAGE_ID", "TYPE_CODE", "AGE", "SEX", "BAND", "UWCLASS", "DURATION", "VALUE")


class LoaderConfig:
    """Externalized, business-maintained inputs. Nothing actuarial is hardcoded."""

    def __init__(self, effdate=S.STANDARD_EFFDATE, isscntry="0000", issuest="00", source_decimals=2):
        self.effdate = effdate          # YYYYMMDD; authoritative single generation (19000101)
        self.isscntry = isscntry        # default country segmentation
        self.issuest = issuest          # default state segmentation
        self.source_decimals = source_decimals

    @classmethod
    def from_dict(cls, d):
        d = d or {}
        return cls(effdate=str(d.get("effdate", S.STANDARD_EFFDATE)),
                   isscntry=str(d.get("isscntry", "0000")),
                   issuest=str(d.get("issuest", "00")),
                   source_decimals=int(d.get("source_decimals", 2)))


def load_plan_crosswalk(xlsx_path):
    """Return (cov2plan, plan2desc) from the authoritative Policy Form Crosswalk (col A -> col C)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cov2plan, plan2desc = {}, {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        cov = row[0] if len(row) > 0 else None
        plan = row[2] if len(row) > 2 else None
        desc = row[4] if len(row) > 4 else None
        if cov and plan:
            c = str(cov).strip(); p = str(plan).strip()
            cov2plan[c] = p
            plan2desc[p] = str(desc).strip() if desc else ""
    return cov2plan, plan2desc


def _to_float(s):
    s = (s or "").strip()
    if s in ("", ".", "-", "-."):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def transform_source(source_csv, cov2plan, config):
    """
    Stream the LifePRO rate extract and classify/transform each row.

    Yields dicts with keys:
      status: IN_SCOPE | EXCLUDED | PLAN_UNRESOLVED | PLAN_INVALID | BAD_VALUE
      and (for IN_SCOPE) the transformed target key + value + provenance.
    """
    with open(source_csv, encoding="utf-8-sig", errors="replace", newline="") as f:
        rd = csv.reader(f)
        next(rd, None)  # header
        lineno = 1
        for r in rd:
            lineno += 1
            if len(r) < 8:
                continue
            cov = r[0].strip(); typ = r[1].strip()
            if cov and set(cov) == {"-"}:
                continue  # dashed separator
            age = r[2].strip(); sex = r[3].strip(); band = r[4].strip()
            uw = r[5].strip(); dur = r[6].strip(); val = r[7].strip()

            if typ in S.EXCLUDED_TYPE_CODES:
                yield {"status": "EXCLUDED", "type_code": typ, "coverage_id": cov, "lineno": lineno}
                continue
            table = S.TYPE_TO_TABLE.get(typ)
            if table is None:
                yield {"status": "EXCLUDED", "type_code": typ, "coverage_id": cov,
                       "lineno": lineno, "note": "unmapped TYPE_CODE"}
                continue

            plan = cov2plan.get(cov)
            if not plan:
                yield {"status": "PLAN_UNRESOLVED", "type_code": typ, "coverage_id": cov,
                       "lineno": lineno}
                continue
            if " " in plan or not plan:
                yield {"status": "PLAN_INVALID", "type_code": typ, "coverage_id": cov,
                       "plan": plan, "lineno": lineno}
                continue

            value = _to_float(val)
            if value is None:
                yield {"status": "BAD_VALUE", "type_code": typ, "coverage_id": cov,
                       "plan": plan, "raw_value": val, "lineno": lineno}
                continue
            try:
                ql_dur = S.source_duration_to_ql(dur)
            except ValueError:
                yield {"status": "BAD_VALUE", "type_code": typ, "coverage_id": cov,
                       "plan": plan, "raw_duration": dur, "lineno": lineno}
                continue
            if ql_dur < 0:
                yield {"status": "BAD_VALUE", "type_code": typ, "coverage_id": cov,
                       "plan": plan, "raw_duration": dur, "lineno": lineno}
                continue

            gender = S.map_sex(sex)
            uwclass = S.map_uwclass(uw)
            band2 = S.map_band(band)
            cntl, col = S.duration_to_cntl_col(ql_dur)

            # Controlled AGE capping (business rule): QLAdmin AGE is C2 (0-99). Source ages
            # above 99 are capped to 99 with preserved lineage. Magnitude of the cap is audited;
            # rows are never discarded and emit is never blocked by the cap itself.
            original_age = age
            emitted_age_int = age.zfill(2)
            age_capped = False
            if age.isdigit() and int(age) > S.MAX_AGE:
                emitted_age_int = str(S.MAX_AGE).zfill(2)
                age_capped = True
            age2 = emitted_age_int

            yield {
                "status": "IN_SCOPE",
                "coverage_id": cov, "type_code": typ, "table": table,
                "plan": plan, "age": age2, "cntl": cntl, "col": col,
                "gender": gender, "uwclass": uwclass, "band": band2,
                "isscntry": config.isscntry, "issuest": config.issuest, "effdate": config.effdate,
                "source_duration": dur, "ql_duration": ql_dur,
                "value": value, "raw_value": val, "lineno": lineno,
                "original_age": original_age, "age_capped": age_capped,
            }


def build_factor_grid(transformed_iter, config):
    """
    Pivot IN_SCOPE transformed cells into factor rows.

    Cell collision precedence (so AGE capping never corrupts genuine data):
      * genuine (non-capped) source data ALWAYS wins over a capped row;
      * a capped cell colliding with genuine data is dropped and recorded in cap_collisions
        (audited, never silent);
      * two genuine cells colliding is a real duplicate -> recorded in collisions (BLOCKER).

    Returns:
      grids: {table: {key_tuple: {col: (value, raw_value, lineno, capped)}}}
      collisions: genuine duplicate cells (table, key, col, prior_lineno, lineno)
      cap_collisions: cap-induced collisions resolved in favor of genuine data
                      (table, key, col, plan, type_code, dropped_value, kept_value)
    key_tuple = (PLAN, AGE, CNTL, GENDER, UWCLASS, BAND, ISSCNTRY, ISSUEST, EFFDATE)
    """
    grids = collections.defaultdict(lambda: collections.defaultdict(dict))
    collisions = []
    cap_collisions = []
    for t in transformed_iter:
        if t.get("status") != "IN_SCOPE":
            continue
        key = (t["plan"], t["age"], t["cntl"], t["gender"], t["uwclass"],
               t["band"], t["isscntry"], t["issuest"], t["effdate"])
        cell = grids[t["table"]][key]
        col = t["col"]
        cur = (t["value"], t["raw_value"], t["lineno"], t.get("age_capped", False),
               t.get("segment_tier", 0))
        prior = cell.get(col)
        if prior is None:
            cell[col] = cur
            continue
        prior_capped = prior[3]
        cur_capped = cur[3]
        if prior_capped and not cur_capped:
            cap_collisions.append((t["table"], key, col, t["plan"], t["type_code"],
                                   prior[0], cur[0]))  # capped prior dropped for genuine
            cell[col] = cur
        elif (not prior_capped) and cur_capped:
            cap_collisions.append((t["table"], key, col, t["plan"], t["type_code"],
                                   cur[0], prior[0]))  # capped current dropped for genuine
        elif prior_capped and cur_capped:
            cap_collisions.append((t["table"], key, col, t["plan"], t["type_code"],
                                   cur[0], prior[0]))  # keep first capped
        elif t.get("source") == "PAAGERAT":
            prior_tier = prior[4] if len(prior) > 4 else 0
            cur_tier = cur[4]
            if cur_tier < prior_tier:
                cell[col] = cur
            elif cur_tier > prior_tier:
                pass
            elif round(prior[0], 8) == round(cur[0], 8):
                pass
            else:
                # sibling PAAGERAT segments under the same parent: first row wins (stream order)
                pass
        else:
            collisions.append((t["table"], key, col, prior[2], t["lineno"]))  # genuine dup
            cell[col] = cur
    return grids, collisions, cap_collisions


def grid_to_factor_rows(table, grid, config):
    """
    Materialize factor rows for one table. Each row is an ordered dict matching
    factor_table_fields(table). Factor cells are formatted via S.format_factor.

    Returns (rows, fmt_issues) where fmt_issues lists cells that could not fit or lost precision.
    """
    pfx = S.PREFIX[table]
    rows = []
    fmt_issues = []
    for key, cells in grid.items():
        plan, age, cntl, gender, uwclass, band, isscntry, issuest, effdate = key
        row = {"PLAN": plan, "AGE": age, "CNTL": cntl, "GENDER": gender, "UWCLASS": uwclass,
               "BAND": band, "ISSCNTRY": isscntry, "ISSUEST": issuest, "EFFDATE": effdate}
        for i in range(S.N_DURATION_COLS):
            field = f"{pfx}{i}"
            if i in cells:
                value, raw_value, lineno = cells[i][0], cells[i][1], cells[i][2]
                text, fits, reduced = S.format_factor(
                    value, max_len=S.factor_field_len(table), source_decimals=config.source_decimals)
                row[field] = text
                if not fits:
                    fmt_issues.append({"table": table, "key": key, "field": field,
                                       "value": value, "issue": "DOES_NOT_FIT", "lineno": lineno})
                elif reduced:
                    fmt_issues.append({"table": table, "key": key, "field": field,
                                       "value": value, "text": text,
                                       "issue": "PRECISION_REDUCED", "lineno": lineno})
            else:
                row[field] = ""  # unpopulated duration cell
        rows.append(row)
    return rows, fmt_issues


# ---- ISWL Phase 1 (Issue #31) — QUIKCVS validation helpers (no emit filter) ----
ISWL_COVERAGE_IDS = frozenset({
    "658 CEN I", "658 CEN SD", "659 CEN II", "659 CEN SR", "659 CEN SD",
    "659 SR GD", "669 SR GD", "679 CEN SD",
})


def quikcvs_keys_by_plan(grid):
    """Return {PLAN: count} of distinct QuikCvs grid keys. grid from build_factor_grid."""
    counts = {}
    for key in grid:
        plan = key[0]
        counts[plan] = counts.get(plan, 0) + 1
    return counts


def quikgps_keys_by_plan(grid):
    """Return {PLAN: count} of distinct QuikGps grid keys."""
    counts = {}
    for key in grid:
        plan = key[0]
        counts[plan] = counts.get(plan, 0) + 1
    return counts


def quikcoi_keys_by_plan(grid):
    """Return {PLAN: count} of distinct QuikCoi grid keys."""
    counts = {}
    for key in grid:
        plan = key[0]
        counts[plan] = counts.get(plan, 0) + 1
    return counts


def quikgcoi_keys_by_plan(grid):
    """Return {PLAN: count} of distinct QuikGcoi grid keys."""
    counts = {}
    for key in grid:
        plan = key[0]
        counts[plan] = counts.get(plan, 0) + 1
    return counts
