"""Build CFIC rate load tracker CSV from Citizens Excel sources (regenerate after catalog updates)."""
from __future__ import annotations

import csv
import re
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit("pip install openpyxl") from exc

ROOT = Path(__file__).resolve().parents[1]
TRACKING = ROOT / "tracking"
CATALOG = ROOT / "Citizens_Plan_Rate_Requirements_Catalog.xlsx"
CROSSWALK = ROOT / "Citizens_Plan_Crosswak.xlsx"

# Catalog column -> QLAdmin tables (factor + key where applicable)
RATE_MAP = [
    ("Gross Premium", "QuikGps", "QuikPlGp", "Access / premium tables"),
    ("Cash / Surrender Values", "QuikCvs", "QuikPlCv", "Green sheets / CFIC_Cash_Values"),
    ("Net Premium", "QuikNps", "QuikPlTv", "Green sheets (paid-up)"),
    ("Reserve / Valuation", "QuikTvs", "QuikPlTv", "Green sheets (terminal reserve)"),
    ("Dividend / PUA", "QuikDvs", "QuikPlDv", "Citizens dividend tables"),
    ("Loan Values", "", "", "Loan factor tables / plan setup"),
    ("Loan Interest Rate", "", "", "Plan setup / QuikPlan fields"),
    ("Cost of Insurance", "QuikCoi", "", "UL/term COI tables"),
    ("Expense Charges", "", "", "Plan setup / expense segments"),
    ("Interest Crediting", "", "", "Plan setup / crediting tables"),
    ("Other Required Factors", "QuikNff", "", "NFO / ETI / special factors"),
]

SKIP_PREFIXES = (
    "not expected",
    "not applicable",
    "not indicated",
    "not stated",
    "no current billing rate",
    "no premium rate expected",
)

SKIP_EXACT = {
    "statutory reserve only",
    "life net premium not applicable",
}

PRIORITY_FAMILIES = {
    "Traditional Permanent Life": "Wave 1",
    "Term Life": "Wave 2",
    "Rider": "Wave 3",
    "Paid-Up / Nonforfeiture": "Wave 2",
    "Nonforfeiture Status": "Wave 2",
}

TRACKER_FIELDS = [
    "tracker_id",
    "cfic_plan_code",
    "ql_plan",
    "in_plan_crosswalk",
    "product_family",
    "product_form",
    "rate_category",
    "requirement_level",
    "confidence",
    "qladmin_factor_table",
    "qladmin_key_table",
    "expected_source",
    "priority_wave",
    "source_received",
    "source_received_date",
    "source_file_location",
    "extract_status",
    "extract_complete_date",
    "load_package_ready",
    "qladmin_loaded",
    "qladmin_load_date",
    "loaded_by",
    "validated_in_qladmin",
    "validation_date",
    "inventory_status_catalog",
    "rate_gap_decision_catalog",
    "notes",
    "last_updated",
]


def load_crosswalk() -> dict[str, dict]:
    wb = openpyxl.load_workbook(CROSSWALK, data_only=True)
    ws = wb["Sheet1"]
    lookup: dict[str, dict] = {}
    for lob, plan, suffix, ql in ws.iter_rows(min_row=2, values_only=True):
        if plan is None:
            continue
        group = str(plan).strip()
        ql_plan = "" if ql is None else str(ql).strip()
        lob_s = "" if lob is None else str(lob).strip()
        for token in re.split(r",\s*", group):
            token = token.strip()
            if token:
                lookup[token.upper()] = {
                    "cfic_plan_group": group,
                    "ql_plan": ql_plan,
                    "lob": lob_s,
                    "suffix": suffix,
                }
    return lookup


def norm_requirement(value: object) -> str:
    return "" if value is None else str(value).strip()


def requirement_in_scope(level: str, rate_category: str) -> bool:
    if not level:
        return False
    low = level.lower()
    if low in SKIP_EXACT:
        return False
    if any(low.startswith(prefix) for prefix in SKIP_PREFIXES):
        return False
    # Loan interest column stores plan rates (e.g. 8%) rather than Expected/Not expected.
    if rate_category == "Loan Interest Rate" and "%" in level:
        return True
    return True


def build_rows() -> list[dict[str, str]]:
    crosswalk = load_crosswalk()
    wb = openpyxl.load_workbook(CATALOG, data_only=True)
    ws = wb["Plan Rate Matrix"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {h: i for i, h in enumerate(headers) if h}

    rows: list[dict[str, str]] = []
    today = date.today().isoformat()

    for record in ws.iter_rows(min_row=2, values_only=True):
        plan = record[col_index["Plan Code"]]
        if not plan:
            continue
        plan_code = str(plan).strip()
        family = norm_requirement(record[col_index["Product Family"]])
        form = norm_requirement(record[col_index["Product Form"]])
        confidence = norm_requirement(record[col_index["Confidence"]])
        inv_status = norm_requirement(record[col_index["Inventory Status"]])
        gap_decision = norm_requirement(record[col_index["Rate Gap Decision"]])
        xwalk = crosswalk.get(plan_code.upper(), {})
        ql_plan = xwalk.get("ql_plan", "")
        in_xwalk = "Y" if xwalk else "N"
        wave = PRIORITY_FAMILIES.get(family, "Later")

        for rate_category, factor, key, source in RATE_MAP:
            req = norm_requirement(record[col_index[rate_category]])
            if not requirement_in_scope(req, rate_category):
                continue
            tracker_id = f"{plan_code}|{rate_category}"
            rows.append(
                {
                    "tracker_id": tracker_id,
                    "cfic_plan_code": plan_code,
                    "ql_plan": ql_plan,
                    "in_plan_crosswalk": in_xwalk,
                    "product_family": family,
                    "product_form": form,
                    "rate_category": rate_category,
                    "requirement_level": req,
                    "confidence": confidence,
                    "qladmin_factor_table": factor,
                    "qladmin_key_table": key,
                    "expected_source": source,
                    "priority_wave": wave,
                    "source_received": "N",
                    "source_received_date": "",
                    "source_file_location": "",
                    "extract_status": "Not Started",
                    "extract_complete_date": "",
                    "load_package_ready": "N",
                    "qladmin_loaded": "N",
                    "qladmin_load_date": "",
                    "loaded_by": "",
                    "validated_in_qladmin": "N",
                    "validation_date": "",
                    "inventory_status_catalog": inv_status,
                    "rate_gap_decision_catalog": gap_decision,
                    "notes": "",
                    "last_updated": today,
                }
            )
    rows.sort(key=lambda r: (r["priority_wave"], r["cfic_plan_code"], r["rate_category"]))
    return rows


OVERRIDE_FIELDS = [
    "source_received",
    "source_received_date",
    "source_file_location",
    "extract_status",
    "extract_complete_date",
    "load_package_ready",
    "qladmin_loaded",
    "qladmin_load_date",
    "loaded_by",
    "validated_in_qladmin",
    "validation_date",
    "notes",
    "last_updated",
]


def load_overrides(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return {row["tracker_id"]: row for row in reader if row.get("tracker_id")}


def apply_overrides(rows: list[dict[str, str]], overrides: dict[str, dict[str, str]]) -> None:
    for row in rows:
        override = overrides.get(row["tracker_id"])
        if not override:
            continue
        for field in OVERRIDE_FIELDS:
            value = override.get(field)
            if value is not None and str(value).strip() != "":
                row[field] = str(value).strip()


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=TRACKER_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def build_summary(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    from collections import defaultdict

    by_wave: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for row in rows:
        wave = row["priority_wave"]
        by_wave[wave]["tracker_rows"] += 1
        if row["qladmin_loaded"] == "Y":
            by_wave[wave]["loaded"] += 1
        if row["source_received"] == "Y":
            by_wave[wave]["source_received"] += 1
        if row["extract_status"] == "Complete":
            by_wave[wave]["extract_complete"] += 1

    summary = []
    for wave in sorted(by_wave):
        d = by_wave[wave]
        summary.append(
            {
                "priority_wave": wave,
                "tracker_rows": d["tracker_rows"],
                "source_received": d.get("source_received", 0),
                "extract_complete": d.get("extract_complete", 0),
                "qladmin_loaded": d.get("loaded", 0),
                "pct_loaded": f"{100 * d.get('loaded', 0) / d['tracker_rows']:.1f}%",
            }
        )
    return summary


def main() -> None:
    rows = build_rows()
    override_path = TRACKING / "CFIC_Rate_Load_Status_Overrides.csv"
    apply_overrides(rows, load_overrides(override_path))
    tracker_path = TRACKING / "CFIC_Rate_Load_Tracker.csv"
    wave1_path = TRACKING / "CFIC_Rate_Load_Tracker_Wave1.csv"
    summary_path = TRACKING / "CFIC_Rate_Load_Summary.csv"

    write_csv(tracker_path, rows)
    wave1 = [r for r in rows if r["priority_wave"] == "Wave 1"]
    write_csv(wave1_path, wave1)

    summary = build_summary(rows)
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["priority_wave", "tracker_rows", "source_received", "extract_complete", "qladmin_loaded", "pct_loaded"],
        )
        writer.writeheader()
        writer.writerows(summary)

    print(f"Tracker rows: {len(rows)}")
    print(f"Wave 1 rows: {len(wave1)}")
    print(f"Wrote: {tracker_path}")
    print(f"Wrote: {wave1_path}")
    print(f"Wrote: {summary_path}")


if __name__ == "__main__":
    main()
