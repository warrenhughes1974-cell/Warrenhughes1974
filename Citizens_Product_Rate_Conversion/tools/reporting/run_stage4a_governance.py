#!/usr/bin/env python3
"""
Stage 4A — Source authority + plan universe reconciliation (governance only).
Writes only under Citizens_Product_Rate_Conversion reports/governance, manifests, docs.
Never modifies CFIC_Rates. Never runs conversion. Never installs engine.
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

CITIZENS = Path(r"C:\Users\warren\Documents\GitHub\Warrenhughes1974\Citizens_Product_Rate_Conversion")
GOV = CITIZENS / "reports" / "governance"


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def norm_code(code: str) -> str:
    s = (code or "").strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def split_codes(cell: str) -> list[str]:
    if cell is None:
        return []
    parts = re.split(r"[,;/]", str(cell))
    out = []
    for p in parts:
        p = p.strip()
        if p and re.match(r"^[A-Za-z0-9.$%\-]{1,12}$", p):
            out.append(p)
    return out


# ---------------------------------------------------------------------------
# Load sources
# ---------------------------------------------------------------------------

def load_tracker():
    path = CITIZENS / "discovery/rates/CFIC_Rate_Load_Tracker.csv"
    rows = list(csv.DictReader(path.open(encoding="utf-8")))
    by_plan = defaultdict(list)
    for r in rows:
        by_plan[r["cfic_plan_code"]].append(r)
    return rows, dict(by_plan)


def load_plans_dbf():
    path = CITIZENS / "staging/normalized_plans/staging/plans_master.csv"
    rows = list(csv.DictReader(path.open(encoding="utf-8")))
    by_plan = {r["pl_plan"]: r for r in rows}
    return rows, by_plan


def load_crosswalk():
    path = CITIZENS / "mappings/working/plans/Citizens_Plan_Crosswalk.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    code_to_ql = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[1] is None:
            continue
        lob, cfic, suffix, ql = row[0], row[1], row[2], row[3]
        codes = split_codes(str(cfic))
        rows.append({"lob": lob, "cfic_raw": cfic, "codes": codes, "suffix": suffix, "qlplan": ql})
        for c in codes:
            code_to_ql[c] = {"qlplan": ql, "lob": lob, "grouped_with": codes, "raw": str(cfic)}
    wb.close()
    return rows, code_to_ql


def load_rate_requirements():
    path = CITIZENS / "mappings/working/rate_types/Citizens_Plan_Rate_Requirements_Catalog.xlsx"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Plan Rate Matrix"]
    headers = [h for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    by_plan = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rec = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
        by_plan[str(row[0]).strip()] = rec
    wb.close()
    return by_plan


def load_reserve_plans():
    root = CITIZENS / "staging/normalized_rates/reserve/staging"
    return {d.name for d in root.iterdir() if d.is_dir()} if root.exists() else set()


def load_draft_plans():
    emit = CITIZENS / "reports/audit/emit_summary.json"
    if emit.exists():
        return set(str(p) for p in json.loads(emit.read_text(encoding="utf-8")).get("plans", []))
    return set()


def load_cash_value_families():
    root = CITIZENS / "source/original/cash_values"
    fams = set()
    if root.exists():
        for z in root.glob("*_CV.zip"):
            fams.add(z.name.replace("_CV.zip", ""))
    return fams


def load_access_plan_hints():
    """Access tables are product-family level, not CFIC plan codes — mark family presence."""
    access = CITIZENS / "source/extracts/access"
    names = {p.stem for p in access.glob("*.csv")} if access.exists() else set()
    return names


def classify_base_or_rider(code: str, tracker_rows: list, req: dict | None) -> str:
    if req:
        fam = str(req.get("Product Family") or "")
        form = str(req.get("Product Form") or "")
        if "rider" in fam.lower() or "rider" in form.lower() or "supplemental" in form.lower():
            return "RIDER"
    if tracker_rows:
        fam = (tracker_rows[0].get("product_family") or "").lower()
        form = (tracker_rows[0].get("product_form") or "").lower()
        if "rider" in fam or "rider" in form or "supplemental" in form:
            return "RIDER"
    # heuristic suffixes often encode sex/smoker for base WL
    return "BASE"


def sex_smoker_family(code: str) -> str | None:
    """Detect P7MN/P7FN/P7MS/P7FS style quads."""
    m = re.match(r"^([A-Z0-9]+?)([MF])([NSJ])$", code.upper())
    if m:
        return m.group(1)
    return None


def build_reconciliation():
    tracker_rows, tracker_by = load_tracker()
    dbf_rows, dbf_by = load_plans_dbf()
    xwalk_rows, xwalk_by = load_crosswalk()
    req_by = load_rate_requirements()
    reserve = load_reserve_plans()
    draft = load_draft_plans()
    cv_fams = load_cash_value_families()
    access_tables = load_access_plan_hints()

    all_raw = set()
    all_raw |= set(tracker_by)
    all_raw |= set(dbf_by)
    all_raw |= set(xwalk_by)
    all_raw |= set(req_by)
    all_raw |= reserve
    all_raw |= draft

    # Also include codes only in grouped crosswalk that might already be there
    records = []
    rid = 0
    for code in sorted(all_raw, key=lambda x: (x.upper(), x)):
        rid += 1
        ncode = norm_code(code)
        trows = tracker_by.get(code, [])
        dbf = dbf_by.get(code)
        xw = xwalk_by.get(code)
        req = req_by.get(code)
        in_t = code in tracker_by
        in_d = code in dbf_by
        in_c = code in xwalk_by
        in_rr = code in req_by
        in_rs = code in reserve
        in_dq = code in draft

        # cash value archive family match (P7MN -> P7MN zip exists)
        in_cv = "Y" if code in cv_fams or any(code.startswith(f) or f.startswith(code[:3]) for f in cv_fams if len(f) >= 3 and code.upper().startswith(f.upper()[:3])) else "N"
        # tighten: exact family zip name
        in_cv = "Y" if code in cv_fams else "N"
        # also mark if prefix family like P7MN in P7MN_CV
        for f in cv_fams:
            if code.upper() == f.upper() or code.upper().startswith(f.upper()):
                in_cv = "Y"

        name = ""
        family = ""
        if dbf:
            name = (dbf.get("pl_desc") or "").strip()
        if trows:
            family = trows[0].get("product_family") or ""
            if not name:
                name = family
        if req:
            family = family or (req.get("Product Family") or "")
            if not name:
                name = family

        base_rider = classify_base_or_rider(code, trows, req)
        fam_key = sex_smoker_family(ncode)

        # scope proposal
        if in_t or in_d or in_rr:
            if base_rider == "RIDER":
                scope = "IN_SCOPE_RIDER"
            else:
                scope = "IN_SCOPE_BASE_PLAN"
        elif in_rs or in_dq:
            scope = "IN_SCOPE_PENDING_SOURCE"
        elif in_c and not in_t and not in_d:
            scope = "REQUIRES_REVIEW"
        else:
            scope = "UNKNOWN"

        # reconciliation status
        if in_t and in_d:
            status = "FULLY_MATCHED"
        elif in_t and not in_d:
            status = "TRACKER_ONLY"
        elif in_d and not in_t:
            status = "DBF_ONLY"
        elif in_c and not in_t and not in_d:
            status = "CROSSWALK_ONLY"
        elif in_rs and not in_t and not in_d:
            status = "RESERVE_STAGING_ONLY"
        else:
            status = "REQUIRES_INTERNAL_REVIEW"

        if in_t and not in_c and not (trows and trows[0].get("ql_plan")):
            if status == "FULLY_MATCHED":
                pass
            mapping_note = "MISSING_MAPPING"
        else:
            mapping_note = ""

        if in_t and not in_c and not (trows and (trows[0].get("ql_plan") or "").strip()):
            if not in_c:
                # keep status but flag
                pass

        has_mapping = "Y" if in_c or (trows and (trows[0].get("ql_plan") or "").strip()) else "N"
        has_rate_req = "Y" if in_t or in_rr else "N"
        has_rate_src = "Y" if in_rs or in_cv == "Y" else "N"

        if has_mapping == "N" and (in_t or in_d):
            recon_extra = "MISSING_MAPPING"
        else:
            recon_extra = status

        # refine status when missing mapping
        final_status = status
        if status == "FULLY_MATCHED" and has_mapping == "N":
            final_status = "MISSING_MAPPING"
        if status == "TRACKER_ONLY" and has_mapping == "N":
            final_status = "TRACKER_ONLY"  # already implies review

        possible_alias = ""
        if fam_key:
            possible_alias = f"SEX_SMOKER_FAMILY:{fam_key}"

        # active/historical - unknown without client
        active = "UNKNOWN"

        req_decision = ""
        if status == "TRACKER_ONLY":
            req_decision = "CIT-DEC-002"
        elif status == "DBF_ONLY":
            req_decision = "CIT-DEC-003"
        elif has_mapping == "N" and (in_t or in_d):
            req_decision = "CIT-DEC-004"

        confidence = "HIGH" if status == "FULLY_MATCHED" else ("MEDIUM" if in_t or in_d else "LOW")

        records.append({
            "RECONCILIATION_ID": f"CIT-PR-{rid:04d}",
            "SOURCE_PLAN_CODE": code,
            "NORMALIZED_PLAN_CODE": ncode,
            "SOURCE_PLAN_NAME": name,
            "NORMALIZED_PLAN_NAME": name,
            "PRODUCT_FAMILY": family,
            "BASE_OR_RIDER": base_rider,
            "ACTIVE_OR_HISTORICAL": active,
            "ISSUE_DATE_FROM": "",
            "ISSUE_DATE_TO": "",
            "IN_TRACKER": "Y" if in_t else "N",
            "IN_PLAN_DBF": "Y" if in_d else "N",
            "IN_CROSSWALK": "Y" if in_c else "N",
            "IN_RATE_REQUIREMENTS": "Y" if in_rr else "N",
            "IN_RESERVE_STAGING": "Y" if in_rs else "N",
            "IN_DRAFT_QUIK_OUTPUT": "Y" if in_dq else "N",
            "IN_ACCESS_EXTRACT": "Y" if any(a.lower() in (family or "").lower().replace(" ", "") for a in access_tables) else "N",
            "IN_PRODUCT_DOCUMENTS": "N",
            "IN_CASH_VALUE_ARCHIVES": in_cv,
            "IN_GROSS_PREMIUM_SOURCE": "Y" if family and any(x in family.lower() for x in ("permalife", "term", "quest", "lpi")) else "N",
            "HAS_POLICY_POPULATION_EVIDENCE": "UNKNOWN",
            "HAS_RATE_REQUIREMENT": has_rate_req,
            "HAS_RATE_SOURCE": has_rate_src,
            "HAS_QLADMIN_MAPPING": has_mapping,
            "POSSIBLE_ALIAS": possible_alias,
            "POSSIBLE_PARENT_PLAN": fam_key or "",
            "POSSIBLE_RIDER_RELATIONSHIP": "Y" if base_rider == "RIDER" else "N",
            "POSSIBLE_STATE_VARIATION": "UNKNOWN",
            "POSSIBLE_ISSUE_ERA_VARIATION": "UNKNOWN",
            "POSSIBLE_DUPLICATE": "N",
            "POSSIBLE_TYPO": "N",
            "PROPOSED_SCOPE_STATUS": scope,
            "PROPOSED_AUTHORITY_STATUS": "PENDING_REVIEW",
            "RECONCILIATION_STATUS": final_status,
            "CONFIDENCE": confidence,
            "REQUIRED_DECISION": req_decision,
            "RELATED_DECISION_ID": req_decision,
            "NOTES": "",
        })

    meta = {
        "tracker_plans": len(tracker_by),
        "dbf_plans": len(dbf_by),
        "crosswalk_codes": len(xwalk_by),
        "crosswalk_rows": len(xwalk_rows),
        "rate_req_plans": len(req_by),
        "reserve_plans": len(reserve),
        "draft_plans": len(draft),
        "cv_families": sorted(cv_fams),
        "raw_distinct": len(all_raw),
        "tracker_only": sorted(set(tracker_by) - set(dbf_by)),
        "dbf_only": sorted(set(dbf_by) - set(tracker_by)),
        "tracker_by": tracker_by,
        "dbf_by": dbf_by,
        "xwalk_by": xwalk_by,
        "req_by": req_by,
        "reserve": reserve,
        "draft": draft,
        "xwalk_rows": xwalk_rows,
    }
    return records, meta


def build_count_bridge(meta):
    """Explain 308 tracker vs 301 DBF without hiding residuals."""
    tracker = set(meta["tracker_by"])
    dbf = set(meta["dbf_by"])
    both = tracker & dbf
    t_only = sorted(tracker - dbf)
    d_only = sorted(dbf - tracker)

    # Normalize comparison for format variations
    t_norm = {norm_code(c): c for c in tracker}
    d_norm = {norm_code(c): c for c in dbf}
    format_matches = []
    for nc, tc in t_norm.items():
        if nc in d_norm and tc not in dbf and d_norm[nc] not in tracker:
            format_matches.append((tc, d_norm[nc]))
        elif nc in d_norm and tc != d_norm[nc] and tc in tracker and d_norm[nc] in dbf:
            # same normalized different raw - only if both exist as different strings
            if tc != d_norm[nc]:
                format_matches.append((tc, d_norm[nc]))

    rows = []
    step = 0

    def add(desc, delta, running, codes=""):
        nonlocal step
        step += 1
        rows.append({
            "STEP": step,
            "DESCRIPTION": desc,
            "DELTA": delta,
            "RUNNING_COUNT": running,
            "PLAN_CODES": codes,
        })

    running = len(tracker)
    add("Start: distinct tracker plan codes", 0, running, "")
    add("Intersect tracker ∩ DBF (unchanged)", 0, running, f"count={len(both)}")
    add("Subtract tracker-only codes", -len(t_only), running - len(t_only), ";".join(t_only))
    running = running - len(t_only)
    add("Add DBF-only codes", len(d_only), running + len(d_only), ";".join(d_only))
    running = running + len(d_only)
    add("End: distinct DBF plan codes (should equal 301)", 0, running, f"dbf={len(dbf)}")
    add("CHECK: tracker - tracker_only + dbf_only == dbf?", 0,
        len(tracker) - len(t_only) + len(d_only),
        f"equals_dbf={len(tracker) - len(t_only) + len(d_only) == len(dbf)}")
    add("Unresolved residual (must be 0 if bridge closes)", 
        len(dbf) - (len(tracker) - len(t_only) + len(d_only)),
        abs(len(dbf) - (len(tracker) - len(t_only) + len(d_only))),
        "")
    # detail rows for each tracker-only and dbf-only
    for c in t_only:
        rows.append({
            "STEP": "",
            "DESCRIPTION": f"TRACKER_ONLY detail: {c}",
            "DELTA": -1,
            "RUNNING_COUNT": "",
            "PLAN_CODES": c,
        })
    for c in d_only:
        rows.append({
            "STEP": "",
            "DESCRIPTION": f"DBF_ONLY detail: {c}",
            "DELTA": 1,
            "RUNNING_COUNT": "",
            "PLAN_CODES": c,
        })
    return rows, t_only, d_only, both


def build_aliases(records, meta):
    """Candidate relationships — do not merge."""
    rels = []
    rid = 0
    by_norm = defaultdict(list)
    for r in records:
        by_norm[r["NORMALIZED_PLAN_CODE"]].append(r["SOURCE_PLAN_CODE"])

    for nc, codes in by_norm.items():
        uniq = sorted(set(codes))
        if len(uniq) > 1:
            rid += 1
            rels.append({
                "RELATIONSHIP_ID": f"CIT-REL-{rid:04d}",
                "PLAN_CODE_A": uniq[0],
                "PLAN_CODE_B": uniq[1],
                "RELATIONSHIP_TYPE": "FORMAT_VARIATION",
                "EVIDENCE": f"Same normalized code {nc}; raw variants {uniq}",
                "CONFIDENCE": "HIGH",
                "CANONICAL_CODE_RECOMMENDATION": uniq[0],
                "IMPACT_ON_PLAN_COUNT": "May reduce distinct count if merged after approval",
                "IMPACT_ON_RATE_MAPPING": "Must map all variants or choose canonical",
                "APPROVAL_REQUIRED": "Y",
                "NOTES": "Do not auto-merge",
            })

    # Sex/smoker families
    families = defaultdict(list)
    for r in records:
        fam = sex_smoker_family(r["NORMALIZED_PLAN_CODE"])
        if fam:
            families[fam].append(r["SOURCE_PLAN_CODE"])
    for fam, codes in sorted(families.items()):
        uniq = sorted(set(codes))
        if len(uniq) >= 2:
            for i in range(len(uniq) - 1):
                rid += 1
                rels.append({
                    "RELATIONSHIP_ID": f"CIT-REL-{rid:04d}",
                    "PLAN_CODE_A": uniq[i],
                    "PLAN_CODE_B": uniq[i + 1],
                    "RELATIONSHIP_TYPE": "SEX_VARIATION" if uniq[i][-2] != uniq[i+1][-2] else "SMOKER_VARIATION",
                    "EVIDENCE": f"Shared stem {fam}; codes {uniq}",
                    "CONFIDENCE": "MEDIUM",
                    "CANONICAL_CODE_RECOMMENDATION": "KEEP_SEPARATE — typically distinct rate grids",
                    "IMPACT_ON_PLAN_COUNT": "Do not collapse; count as separate plans",
                    "IMPACT_ON_RATE_MAPPING": "May share QLAdmin plan with sex/smoker members",
                    "APPROVAL_REQUIRED": "Y",
                    "NOTES": "Crosswalk often groups these (e.g. P7FN,P7FS,P7MN,P7MS)",
                })

    # Crosswalk consolidations
    for xw in meta["xwalk_rows"]:
        codes = xw["codes"]
        if len(codes) > 1:
            rid += 1
            rels.append({
                "RELATIONSHIP_ID": f"CIT-REL-{rid:04d}",
                "PLAN_CODE_A": codes[0],
                "PLAN_CODE_B": ",".join(codes[1:]),
                "RELATIONSHIP_TYPE": "QLADMIN_CONSOLIDATION",
                "EVIDENCE": f"Crosswalk row CFIC Plan='{xw['cfic_raw']}' → QLPlan={xw['qlplan']}",
                "CONFIDENCE": "HIGH",
                "CANONICAL_CODE_RECOMMENDATION": f"QLPlan {xw['qlplan']} (source codes remain distinct)",
                "IMPACT_ON_PLAN_COUNT": "Source count unchanged; QLAdmin destinations fewer",
                "IMPACT_ON_RATE_MAPPING": "Multiple CFIC codes → one QLAdmin plan",
                "APPROVAL_REQUIRED": "Y",
                "NOTES": "Working mapping — not approved",
            })

    return rels


def map_req_cell(val: str) -> str:
    if val is None:
        return "APPLICABILITY_UNKNOWN"
    s = str(val).strip().lower()
    if not s or s in ("nan", "none"):
        return "APPLICABILITY_UNKNOWN"
    if "not expected" in s or "not applicable" in s or "not indicated" in s:
        return "NOT_APPLICABLE"
    if "not reviewed" in s:
        return "REQUIRES_CLIENT_REVIEW"
    if "expected" in s or "rider premium" in s or "conditional" in s:
        return "REQUIRED_AUTHORITY_PENDING"
    return "REQUIRES_ACTUARIAL_REVIEW"


RATE_COLS = [
    ("Gross Premium", "GROSS_PREMIUM"),
    ("Cash / Surrender Values", "CASH_VALUE"),
    ("Net Premium", "NET_PREMIUM"),
    ("Reserve / Valuation", "RESERVE"),
    ("Dividend / PUA", "DIVIDEND_PUA"),
    ("Loan Values", "LOAN_VALUES"),
    ("Loan Interest Rate", "LOAN_INTEREST"),
    ("Cost of Insurance", "COI"),
    ("Expense Charges", "EXPENSE"),
    ("Interest Crediting", "INTEREST_CREDITING"),
    ("Other Required Factors", "OTHER"),
]


def build_rate_matrix(records, meta):
    rows = []
    req_by = meta["req_by"]
    reserve = meta["reserve"]
    draft = meta["draft"]
    for r in records:
        code = r["SOURCE_PLAN_CODE"]
        req = req_by.get(code, {})
        row = {
            "SOURCE_PLAN_CODE": code,
            "PRODUCT_FAMILY": r["PRODUCT_FAMILY"],
            "BASE_OR_RIDER": r["BASE_OR_RIDER"],
            "PROPOSED_SCOPE_STATUS": r["PROPOSED_SCOPE_STATUS"],
        }
        for col, key in RATE_COLS:
            cell = map_req_cell(req.get(col) if req else None)
            # enrich with extraction evidence for reserve-related
            if key in ("CASH_VALUE", "NET_PREMIUM", "RESERVE") and code in reserve:
                if cell.startswith("REQUIRED"):
                    cell = "REQUIRED_CONVERTED_DRAFT" if code in draft else "REQUIRED_EXTRACTION_PENDING"
                    if code in draft and key in ("CASH_VALUE", "NET_PREMIUM", "RESERVE"):
                        cell = "REQUIRED_CONVERTED_DRAFT"
            if key == "CASH_VALUE" and r["IN_CASH_VALUE_ARCHIVES"] == "Y" and cell.startswith("REQUIRED"):
                if cell != "REQUIRED_CONVERTED_DRAFT":
                    cell = "REQUIRED_SOURCE_IDENTIFIED"
            row[key] = cell if req else ("APPLICABILITY_UNKNOWN" if r["IN_TRACKER"] == "N" else map_req_cell(None))
            if not req and r["IN_TRACKER"] == "Y":
                # tracker has rate categories - approximate from tracker rows
                row[key] = "REQUIRED_AUTHORITY_PENDING"
        # override from tracker categories if present
        trows = meta["tracker_by"].get(code, [])
        cat_map = {
            "Gross Premium": "GROSS_PREMIUM",
            "Cash / Surrender Values": "CASH_VALUE",
            "Net Premium": "NET_PREMIUM",
            "Reserve / Valuation": "RESERVE",
            "Dividend / PUA": "DIVIDEND_PUA",
            "Loan Values": "LOAN_VALUES",
            "Loan Interest Rate": "LOAN_INTEREST",
            "Cost of Insurance": "COI",
            "Expense Charges": "EXPENSE",
            "Interest Crediting": "INTEREST_CREDITING",
            "Other Required Factors": "OTHER",
        }
        for tr in trows:
            cat = tr.get("rate_category", "")
            key = cat_map.get(cat)
            if not key:
                continue
            level = (tr.get("requirement_level") or "").lower()
            if "not" in level and "expect" in level:
                row[key] = "NOT_APPLICABLE"
            elif code in reserve and key in ("CASH_VALUE", "NET_PREMIUM", "RESERVE"):
                row[key] = "REQUIRED_CONVERTED_DRAFT" if code in draft else "REQUIRED_EXTRACTION_PENDING"
            else:
                row[key] = "REQUIRED_AUTHORITY_PENDING"
        # Extended term etc. always unknown unless noted
        row["EXTENDED_TERM"] = "REQUIRES_ACTUARIAL_REVIEW" if r["BASE_OR_RIDER"] == "BASE" else "NOT_APPLICABLE"
        row["PAID_UP"] = row.get("CASH_VALUE", "APPLICABILITY_UNKNOWN")
        row["MODAL_FACTOR"] = "APPLICABILITY_UNKNOWN"
        row["GUIDELINE_PREMIUM"] = "APPLICABILITY_UNKNOWN"
        row["SETTLEMENT_FACTOR"] = "APPLICABILITY_UNKNOWN"
        rows.append(row)
    return rows


def populate_plan_manifest(records):
    cols = [
        "PLAN_ID", "SOURCE_PLAN_CODE", "SOURCE_PLAN_NAME", "QLADMIN_PLAN_CODE",
        "PRODUCT_FAMILY", "BASE_OR_RIDER", "PARTICIPATING_STATUS", "INTEREST_SENSITIVE_STATUS",
        "ISSUE_DATE_FROM", "ISSUE_DATE_TO",
        "SOURCE_REVIEW_STATUS", "BUSINESS_CLASSIFICATION_STATUS", "PLAN_MAPPING_STATUS",
        "PRODUCT_SETUP_STATUS", "RATE_IDENTIFICATION_STATUS", "CONVERSION_STATUS",
        "VALIDATION_STATUS", "CLIENT_REVIEW_STATUS", "OWNER", "OPEN_ISSUE_COUNT",
        "SOURCE_AUTHORITY", "NOTES",
        "PLAN_UNIVERSE_STATUS", "PLAN_AUTHORITY_STATUS", "PLAN_IDENTITY_CONFIDENCE",
        "TRACKER_INDICATOR", "PLAN_DBF_INDICATOR", "CROSSWALK_INDICATOR",
        "RATE_REQUIREMENTS_INDICATOR", "RESERVE_STAGING_INDICATOR", "DRAFT_OUTPUT_INDICATOR",
        "POSSIBLE_ALIAS", "CANONICAL_PLAN_CODE",
        "CLIENT_REVIEW_REQUIRED", "ACTUARIAL_REVIEW_REQUIRED", "INTERNAL_REVIEW_REQUIRED",
        "RELATED_DECISION_ID",
    ]
    rows = []
    for i, r in enumerate(records, 1):
        mapping_status = "IN_PROGRESS" if r["HAS_QLADMIN_MAPPING"] == "Y" else "NOT_STARTED"
        rate_id = "IN_PROGRESS" if r["HAS_RATE_REQUIREMENT"] == "Y" else "NOT_STARTED"
        if r["HAS_RATE_SOURCE"] == "Y":
            rate_id = "READY_FOR_REVIEW"
        conv = "NOT_STARTED"
        if r["IN_DRAFT_QUIK_OUTPUT"] == "Y":
            conv = "IN_PROGRESS"  # draft only — not complete
        rows.append({
            "PLAN_ID": f"CIT-PLAN-{i:04d}",
            "SOURCE_PLAN_CODE": r["SOURCE_PLAN_CODE"],
            "SOURCE_PLAN_NAME": r["SOURCE_PLAN_NAME"],
            "QLADMIN_PLAN_CODE": "",  # filled if crosswalk - below
            "PRODUCT_FAMILY": r["PRODUCT_FAMILY"],
            "BASE_OR_RIDER": r["BASE_OR_RIDER"],
            "PARTICIPATING_STATUS": "UNKNOWN",
            "INTEREST_SENSITIVE_STATUS": "UNKNOWN",
            "ISSUE_DATE_FROM": "",
            "ISSUE_DATE_TO": "",
            "SOURCE_REVIEW_STATUS": "READY_FOR_REVIEW",
            "BUSINESS_CLASSIFICATION_STATUS": "IN_PROGRESS" if r["PRODUCT_FAMILY"] else "NOT_STARTED",
            "PLAN_MAPPING_STATUS": mapping_status,
            "PRODUCT_SETUP_STATUS": "NOT_STARTED",
            "RATE_IDENTIFICATION_STATUS": rate_id,
            "CONVERSION_STATUS": conv,
            "VALIDATION_STATUS": "NOT_STARTED",
            "CLIENT_REVIEW_STATUS": "NOT_STARTED",
            "OWNER": "",
            "OPEN_ISSUE_COUNT": "1" if r["REQUIRED_DECISION"] else "0",
            "SOURCE_AUTHORITY": "PENDING_REVIEW",
            "NOTES": "Working governance row from Stage 4A; not client-approved",
            "PLAN_UNIVERSE_STATUS": r["PROPOSED_SCOPE_STATUS"],
            "PLAN_AUTHORITY_STATUS": "PENDING_REVIEW",
            "PLAN_IDENTITY_CONFIDENCE": r["CONFIDENCE"],
            "TRACKER_INDICATOR": r["IN_TRACKER"],
            "PLAN_DBF_INDICATOR": r["IN_PLAN_DBF"],
            "CROSSWALK_INDICATOR": r["IN_CROSSWALK"],
            "RATE_REQUIREMENTS_INDICATOR": r["IN_RATE_REQUIREMENTS"],
            "RESERVE_STAGING_INDICATOR": r["IN_RESERVE_STAGING"],
            "DRAFT_OUTPUT_INDICATOR": r["IN_DRAFT_QUIK_OUTPUT"],
            "POSSIBLE_ALIAS": r["POSSIBLE_ALIAS"],
            "CANONICAL_PLAN_CODE": r["NORMALIZED_PLAN_CODE"],
            "CLIENT_REVIEW_REQUIRED": "Y" if r["RECONCILIATION_STATUS"] in ("TRACKER_ONLY", "DBF_ONLY", "CROSSWALK_ONLY") else "N",
            "ACTUARIAL_REVIEW_REQUIRED": "Y" if r["BASE_OR_RIDER"] == "BASE" else "N",
            "INTERNAL_REVIEW_REQUIRED": "Y",
            "RELATED_DECISION_ID": r["RELATED_DECISION_ID"],
        })
    # fill QLAdmin from crosswalk
    _, xwalk_by = load_crosswalk()
    for row in rows:
        xw = xwalk_by.get(row["SOURCE_PLAN_CODE"])
        if xw:
            row["QLADMIN_PLAN_CODE"] = xw.get("qlplan") or ""
    write_csv(CITIZENS / "manifests/plan_manifest.csv", cols, rows)
    return rows


def source_register():
    rows = [
        ("SA-001", "Citizens Plans DBF", "cifi0004.dbf", "source/original/dbf/cifi0004.dbf", "DBF",
         "Plan master extract source", "plan_universe;plan_code;plan_name;fees;loan_interest",
         "301 plans", "loan_interest;policy_fees", "PENDING_REVIEW", 1,
         "Client-provided FoxPro plans table; primary candidate for plan identity", "No issue dates fully confirmed in extract; descriptions often blank"),
        ("SA-002", "Rate Load Tracker", "CFIC_Rate_Load_Tracker.csv", "discovery/rates/CFIC_Rate_Load_Tracker.csv", "CSV_TRACKER",
         "Plan×rate-category requirements tracking", "rate_requirements;plan_universe",
         "308 plans", "11 rate categories", "DERIVED_WORKING", 2,
         "Internal derived from requirements catalog + crosswalk; not client-signed", "May include manually added plans; not provenance for plan existence alone"),
        ("SA-003", "Rate Requirements Catalog", "Citizens_Plan_Rate_Requirements_Catalog.xlsx", "mappings/working/rate_types/Citizens_Plan_Rate_Requirements_Catalog.xlsx", "XLSX",
         "Expected rate types by plan", "rate_requirements;product_family",
         "308 plan codes", "gross,cv,net,reserve,div,loan,coi,expense,interest,other", "DERIVED_WORKING", 2,
         "Working BA catalog; confidence High for many riders", "Inventory Status Not Reviewed; Rate Gap Decision TBD"),
        ("SA-004", "Plan Crosswalk", "Citizens_Plan_Crosswalk.xlsx", "mappings/working/plans/Citizens_Plan_Crosswalk.xlsx", "XLSX",
         "CFIC plan → QLAdmin plan working map", "qladmin_mapping",
         "111 rows / ~156 expanded codes", "N/A", "DERIVED_WORKING", 3,
         "Working mapping; groups sex/smoker codes to one QLPlan", "Incomplete vs 308; not approved"),
        ("SA-005", "Reserve DBF", "cifi0007.DBF", "source/original/dbf/cifi0007.DBF", "DBF",
         "Reserve / CV / net / PU grids", "cash_value;net_premium;terminal_reserve;paid_up",
         "Plans present in file (~138 staged)", "CV, terminal reserve, net, PU", "PENDING_REVIEW", 1,
         "Primary technical candidate for CV/reserve/PU; actuarial confirmation required", "Gross premium absent; OBQ-1 basis open"),
        ("SA-006", "Reserve staging grids", "reserve_grid.csv", "staging/normalized_rates/reserve/staging/", "STAGING",
         "Derived extract from reserve DBF", "cash_value;reserves;net;paid_up",
         "138 plans", "CV/reserve/net/PU", "DERIVED_WORKING", 4,
         "Derived from SA-005; not independent authority", "Wave/subset of DBF content"),
        ("SA-007", "Cash value ZIP archives", "*_CV.zip", "source/original/cash_values/", "ZIP_PDF",
         "Green-sheet PDF archives by family", "cash_value;paid_up;eti_candidate",
         "15 families", "CV grids in PDFs", "SUPPORTING", 2,
         "Client-provided illustrations; OCR path failed; DBF preferred when present", "Not extracted; overlap with MultipleCashValueFiles quarantined"),
        ("SA-008", "Access Proposal Maker extracts", "*.csv", "source/extracts/access/", "ACCESS_EXTRACT",
         "Illustration/premium checkpoints by product family", "gross_premium;validation;illustrations",
         "Product families not CFIC plan codes", "Gross premium + sparse CV", "VALIDATION_ONLY", 2,
         "Useful validation vs reserve DBF for PermaLife7", "Not plan-universe authority; sparse CV columns"),
        ("SA-009", "Access MDB original", "CFIProposalMakerRev2.mdb", "source/original/access/", "MDB",
         "Source of Access extracts", "gross_premium;product_catalog",
         "Product tables", "Premiums/illustrations", "SUPPORTING", 2,
         "Client proposal tool database", "Not full rate grid authority"),
        ("SA-010", "Draft Quik output", "Quik*.csv", "output/csv/draft_pre_migration/", "DRAFT_OUTPUT",
         "Historical draft QLAdmin load package", "validation_evidence",
         "138 plans in emit_summary", "CV/TV/NP factors", "NOT_AUTHORITATIVE", 9,
         "Draft only; OBQ blockers; not production", "Must not be treated as completed conversion"),
        ("SA-011", "Product catalog doc", "product_catalog.md", "docs/source_layout/product_catalog.md", "DOCUMENTATION",
         "Access product family inventory", "product_family",
         "Access families", "N/A", "SUPPORTING", 3,
         "Internal discovery from Access", "States products ACTIVE as of 2026-07-08 — needs reconfirm"),
        ("SA-012", "SourceData 11-18-2024", "SourceData_11-18-2024/", "archive/legacy_cfic_rates/SourceData_11-18-2024/", "ARCHIVE",
         "Legacy dump", "unknown",
         "Unknown", "Unknown", "HISTORICAL", 8,
         "Archived; currency unconfirmed", "Do not use as current authority"),
        ("SA-013", "OCR extracts", "docs/_ocr_extract", "archive/legacy_cfic_rates/ocr_extract/", "OCR",
         "Failed/partial OCR of rate sheets", "gross_premium_candidate",
         "PDF pages", "Unreliable", "NOT_AUTHORITATIVE", 9,
         "Pilot FAIL vs Access for CV", "Do not use as rate authority"),
        ("SA-014", "Annuity DBF", "cifianu1.dbf", "quarantine/sensitive_review/cifianu1.dbf", "DBF",
         "Annuity transactions", "out_of_scope_candidate",
         "N/A life plans", "N/A", "PENDING_REVIEW", 9,
         "Quarantined; life-rate scope undecided", "Sensitive/transactional"),
        ("SA-015", "Rate sheet PDFs", "*.pdf", "source/product_documents/rate_sheets/", "PDF",
         "Printed rate sheets", "gross_premium",
         "Subset of products", "Gross premium candidate", "PENDING_REVIEW", 2,
         "Candidate for gross premium with Access", "Issue 02 pilot only"),
        ("SA-016", "OBQ assumptions template", "cfic_rate_key_assumptions.csv", "mappings/working/business_inputs/", "CSV",
         "QuikPlCv/Tv actuarial assumption placeholders", "rate_key_assumptions",
         "Keys from draft", "MORT/INT fields", "DERIVED_WORKING", 5,
         "Blank placeholders awaiting actuarial", "Blocks QLAdmin load"),
    ]
    out = []
    for r in rows:
        out.append({
            "SOURCE_AUTHORITY_ID": r[0],
            "SOURCE_NAME": r[1],
            "SOURCE_FILENAME": r[2],
            "SOURCE_RELATIVE_PATH": r[3],
            "SOURCE_TYPE": r[4],
            "BUSINESS_PURPOSE": r[5],
            "DATA_DOMAIN": r[6],
            "PLAN_COVERAGE": r[7],
            "RATE_TYPE_COVERAGE": r[8],
            "EFFECTIVE_DATE_FROM": "",
            "EFFECTIVE_DATE_TO": "",
            "SOURCE_OWNER": "PENDING",
            "SOURCE_PROVIDER": "Citizens/CFIC client or internal derivation",
            "ORIGINAL_OR_DERIVED": "ORIGINAL" if r[4] in ("DBF", "MDB", "ZIP_PDF", "PDF") else "DERIVED",
            "CURRENT_OR_HISTORICAL": "HISTORICAL" if r[9] == "HISTORICAL" else "CURRENT",
            "CLIENT_PROVIDED_INDICATOR": "Y" if r[4] in ("DBF", "MDB", "ZIP_PDF", "PDF", "XLSX") and "Catalog" not in r[1] and "Crosswalk" not in r[1] and "Tracker" not in r[1] else "N",
            "ACTUARIAL_INDICATOR": "Y" if "reserve" in r[1].lower() or "assumption" in r[1].lower() or "cash" in r[1].lower() else "N",
            "INTERNAL_WORK_PRODUCT_INDICATOR": "Y" if r[9] in ("DERIVED_WORKING", "NOT_AUTHORITATIVE") else "N",
            "APPROVAL_EVIDENCE": "None — Stage 4A proposed only",
            "AUTHORITY_STATUS": r[9],
            "AUTHORITY_RANK": r[10],
            "CONFLICT_RESOLUTION_RULE": r[11],
            "KNOWN_LIMITATIONS": r[12],
            "RELATED_DECISION_ID": "",
            "NOTES": "",
        })
    # fix CLIENT for crosswalk/catalog - they're working artifacts possibly client-origin
    for o in out:
        if "Crosswalk" in o["SOURCE_NAME"] or "Catalog" in o["SOURCE_NAME"]:
            o["CLIENT_PROVIDED_INDICATOR"] = "UNKNOWN"
            o["INTERNAL_WORK_PRODUCT_INDICATOR"] = "Y"
    return out


def source_by_domain():
    domains = [
        ("Plan universe", "Plans DBF (SA-001) PROPOSED", "Tracker (SA-002) SUPPORTING", "Access catalog SUPPORTING", "308 vs 301", "DBF for existence; tracker for requirements coverage; union for review queue", "MEDIUM", "CIT-DEC-001", "Plan manifest / scope"),
        ("Plan code", "Plans DBF pl_plan", "Tracker cfic_plan_code", "Crosswalk CFIC Plan", "Format/alias variants", "Prefer DBF code when present; keep tracker-only as review", "MEDIUM", "CIT-DEC-001", "All mapping"),
        ("Plan name/description", "Plans DBF pl_desc", "Tracker product_family", "Product catalog", "Many blank DBF desc", "Family from tracker/catalog until names filled", "LOW", "CIT-DEC-001", "Reporting"),
        ("Product family", "Rate requirements catalog", "Tracker", "Access product catalog", "Naming inconsistency", "Catalog/tracker working", "MEDIUM", "CIT-DEC-001", "Rate applicability"),
        ("Base vs rider", "Rate requirements Product Form", "Tracker product_form", "Plan code heuristics", "Misclassified riders", "Catalog first", "MEDIUM", "CIT-DEC-006", "Scope counts"),
        ("Participating status", "UNKNOWN", "", "", "No authoritative field identified", "Requires actuarial/client", "LOW", "CIT-DEC-016", "Dividend rates"),
        ("Interest-sensitive status", "UNKNOWN", "", "", "No authoritative field identified", "Requires actuarial/client", "LOW", "CIT-DEC-015", "Interest rates"),
        ("Issue-date range", "UNKNOWN / partial DBF", "", "", "Missing", "Requires client", "LOW", "CIT-DEC-007", "Effective dating"),
        ("State / sex / smoker / band", "Plan code suffixes + crosswalk grouping", "Reserve member dims in draft", "Access sex/smoker cols", "Encoding not formally documented", "Document encoding; keep codes distinct", "MEDIUM", "CIT-DEC-005", "Rate members"),
        ("Rate requirements", "Requirements catalog + tracker", "", "", "Not Reviewed / TBD gaps", "Working until client confirms", "MEDIUM", "CIT-DEC-008", "Rate manifest"),
        ("Gross premiums", "Access extracts + rate sheet PDFs", "Requirements catalog", "PDF pilot", "No DBF gross grid", "PROPOSED Access primary; PDF supporting", "MEDIUM", "CIT-DEC-009", "QuikGps"),
        ("Cash values", "Reserve DBF cifi0007", "CV ZIP PDFs supporting", "Access illustration checkpoints validation", "OBQ-1 basis", "PROPOSED DBF primary", "HIGH", "CIT-DEC-010", "QuikCvs"),
        ("Net premiums", "Reserve DBF", "Access validation", "", "Basis open", "PROPOSED DBF primary", "HIGH", "CIT-DEC-011", "QuikNps related"),
        ("Terminal reserves", "Reserve DBF", "Draft QuikTvs", "", "OBQ-2 keys", "PROPOSED DBF primary", "HIGH", "CIT-DEC-012", "QuikTvs"),
        ("Mean reserves", "Reserve DBF fields (candidate)", "", "", "Mapping undecided", "Actuarial review", "LOW", "CIT-DEC-012", "Optional Quik"),
        ("Paid-up insurance", "Reserve DBF PUP", "Access PaidUp* validation", "CV PDFs", "", "PROPOSED DBF primary", "HIGH", "CIT-DEC-013", "QuikNps"),
        ("Extended-term insurance", "UNKNOWN / green sheets candidate", "OBQ historical", "", "No confirmed load path", "Actuarial", "LOW", "CIT-DEC-014", "ETI"),
        ("Dividends / div interest", "UNKNOWN", "Requirements catalog indicators", "", "Mostly Not indicated", "Client/actuarial", "LOW", "CIT-DEC-016", "QuikDvs"),
        ("Loan interest", "Plans DBF IR1–IR8", "Requirements catalog", "", "Not emitted", "PROPOSED plans DBF", "MEDIUM", "CIT-DEC-015", "Loan IR"),
        ("Guaranteed/current/credited interest", "UNKNOWN", "Requirements Interest Crediting", "", "Sparse", "Actuarial", "LOW", "CIT-DEC-015", "Interest"),
        ("COI / expenses / loads / surrender / modal", "Requirements catalog only", "", "", "Sources unidentified", "Actuarial/client", "LOW", "CIT-DEC-017", "UL/expense"),
        ("Policy fees", "Plans DBF fee fields", "Requirements", "", "", "PROPOSED plans DBF", "MEDIUM", "CIT-DEC-017", "Fees"),
        ("Rider premiums", "Access rider tables + catalog", "PDF/docs", "", "Not plan-coded in Access", "PROPOSED Access+catalog", "MEDIUM", "CIT-DEC-006", "Rider rates"),
        ("Guideline/MEC/target/settlement", "UNKNOWN", "", "", "No source", "Client", "LOW", "CIT-DEC-017", "Advanced premiums"),
        ("QLAdmin destination mapping", "Working crosswalk", "Tracker ql_plan", "Draft keys", "Incomplete 156 vs 308", "Working only until CIT-DEC-004", "MEDIUM", "CIT-DEC-004", "All emit"),
        ("Validation authority", "Access checkpoints + issue evidence", "Draft outputs NOT authority", "", "P7MN only strong", "Validation-only sources", "MEDIUM", "CIT-DEC-018", "UAT"),
        ("Client UAT authority", "NONE YET", "", "", "No sign-off", "Future client package", "LOW", "CIT-DEC-020", "Release"),
        ("Draft Quik evidentiary status", "NOT_AUTHORITATIVE", "SA-010", "", "Useful regression only", "Historical draft", "HIGH", "CIT-DEC-018", "Regression"),
        ("Historical SourceData", "HISTORICAL_ONLY", "SA-012", "", "Unconfirmed", "Archive", "HIGH", "CIT-DEC-019", "None until reviewed"),
    ]
    out = []
    for d in domains:
        out.append({
            "DATA_DOMAIN": d[0],
            "PROPOSED_PRIMARY_SOURCE": d[1],
            "PROPOSED_SECONDARY_SOURCE": d[2],
            "VALIDATION_SOURCE": d[3],
            "CONFLICTING_SOURCES": d[4],
            "MISSING_AUTHORITY": "Y" if "UNKNOWN" in d[1] else "N",
            "RECOMMENDED_PRECEDENCE": d[5],
            "CONFIDENCE": d[6],
            "REQUIRED_APPROVAL": d[7],
            "BLOCKED_CONVERSION_AREAS": d[8],
        })
    return out


def main():
    GOV.mkdir(parents=True, exist_ok=True)

    # Part 1 input validation
    inputs = {
        "Stage3_Architecture_and_Execution_Readiness_Report.md": (CITIZENS / "Stage3_Architecture_and_Execution_Readiness_Report.md").exists(),
        "plan_universe_reconciliation.csv": (CITIZENS / "reports/architecture/plan_universe_reconciliation.csv").exists(),
        "rate_universe_baseline.csv": (CITIZENS / "reports/architecture/rate_universe_baseline.csv").exists(),
        "source_manifest.csv": (CITIZENS / "manifests/source_manifest.csv").exists(),
        "migration_inventory.csv": (CITIZENS / "manifests/migration_inventory.csv").exists(),
        "tracker": (CITIZENS / "discovery/rates/CFIC_Rate_Load_Tracker.csv").exists(),
        "plans_master": (CITIZENS / "staging/normalized_plans/staging/plans_master.csv").exists(),
        "crosswalk": (CITIZENS / "mappings/working/plans/Citizens_Plan_Crosswalk.xlsx").exists(),
        "rate_requirements": (CITIZENS / "mappings/working/rate_types/Citizens_Plan_Rate_Requirements_Catalog.xlsx").exists(),
        "reserve_staging": (CITIZENS / "staging/normalized_rates/reserve/staging").exists(),
        "draft_quik": (CITIZENS / "output/csv/draft_pre_migration").exists(),
        "product_catalog": (CITIZENS / "docs/source_layout/product_catalog.md").exists(),
        "access_extracts": (CITIZENS / "source/extracts/access").exists(),
    }
    (GOV / "Stage4A_Input_Validation_Report.md").write_text(
        "# Stage 4A Input Validation\n\n**Generated:** " + utc_now() + "\n\n| Input | Present |\n|-------|--------|\n"
        + "\n".join(f"| {k} | {v} |" for k, v in inputs.items())
        + f"\n\n**Result:** {'PASS' if all(inputs.values()) else 'FAIL'}\n",
        encoding="utf-8",
    )
    if not all(inputs.values()):
        print("INPUT VALIDATION FAIL", {k: v for k, v in inputs.items() if not v})
        return 2

    print("Building reconciliation...")
    records, meta = build_reconciliation()
    write_csv(GOV / "plan_universe_master_reconciliation.csv", list(records[0].keys()), records)

    print("Count bridge...")
    bridge, t_only, d_only, both = build_count_bridge(meta)
    write_csv(GOV / "plan_count_bridge_308_to_301.csv", list(bridge[0].keys()), bridge)

    # Crosswalk coverage
    xwalk_codes = set(meta["xwalk_by"])
    tracker = set(meta["tracker_by"])
    dbf = set(meta["dbf_by"])
    xrows = []
    for code in sorted(xwalk_codes, key=str.upper):
        xw = meta["xwalk_by"][code]
        xrows.append({
            "SOURCE_PLAN_CODE": code,
            "IN_TRACKER": "Y" if code in tracker else "N",
            "IN_PLAN_DBF": "Y" if code in dbf else "N",
            "QLPLAN": xw.get("qlplan", ""),
            "GROUPED_WITH": ",".join(xw.get("grouped_with") or []),
            "COVERAGE_INTERPRETATION": "WORKING_PARTIAL_MAP — not full universe",
            "NOTES": "Crosswalk is incomplete subset; grouped sex/smoker codes common",
        })
    # summary row interpretation
    write_csv(GOV / "crosswalk_coverage_analysis.csv", list(xrows[0].keys()), xrows)

    # Reserve coverage
    rrows = []
    for code in sorted(meta["reserve"]):
        r = next((x for x in records if x["SOURCE_PLAN_CODE"] == code), None)
        rrows.append({
            "SOURCE_PLAN_CODE": code,
            "IN_TRACKER": r["IN_TRACKER"] if r else "N",
            "IN_PLAN_DBF": r["IN_PLAN_DBF"] if r else "N",
            "IN_DRAFT_QUIK_OUTPUT": "Y" if code in meta["draft"] else "N",
            "COVERAGE_INTERPRETATION": "RESERVE_WAVE_SUBSET — plans successfully staged from cifi0007 (emit_summary lists same 138)",
            "NOTES": "Not all 301/308 plans have reserve grids; absence ≠ not applicable",
        })
    write_csv(GOV / "reserve_staging_coverage_analysis.csv", list(rrows[0].keys()), rrows)

    print("Aliases...")
    rels = build_aliases(records, meta)
    write_csv(GOV / "plan_alias_and_relationship_candidates.csv", list(rels[0].keys()), rels)

    print("Rate matrix...")
    rmat = build_rate_matrix(records, meta)
    write_csv(GOV / "plan_rate_requirement_authority_matrix.csv", list(rmat[0].keys()), rmat)

    print("Source register / domains...")
    sreg = source_register()
    write_csv(CITIZENS / "manifests/source_authority_register.csv", list(sreg[0].keys()), sreg)
    sdom = source_by_domain()
    write_csv(GOV / "source_authority_by_domain.csv", list(sdom[0].keys()), sdom)

    print("Plan manifest...")
    pm = populate_plan_manifest(records)

    # Stats
    scope_counts = Counter(r["PROPOSED_SCOPE_STATUS"] for r in records)
    status_counts = Counter(r["RECONCILIATION_STATUS"] for r in records)
    base_n = sum(1 for r in records if r["PROPOSED_SCOPE_STATUS"] == "IN_SCOPE_BASE_PLAN")
    rider_n = sum(1 for r in records if r["PROPOSED_SCOPE_STATUS"] == "IN_SCOPE_RIDER")
    missing_map = sum(1 for r in records if r["HAS_QLADMIN_MAPPING"] == "N" and r["PROPOSED_SCOPE_STATUS"].startswith("IN_SCOPE"))
    missing_src = sum(1 for r in records if r["HAS_RATE_SOURCE"] == "N" and r["PROPOSED_SCOPE_STATUS"].startswith("IN_SCOPE"))
    unresolved = sum(1 for r in records if r["PROPOSED_SCOPE_STATUS"] in ("UNKNOWN", "REQUIRES_REVIEW", "IN_SCOPE_PENDING_SOURCE"))
    norm_codes = len({r["NORMALIZED_PLAN_CODE"] for r in records})

    summary = {
        "timestamp_utc": utc_now(),
        "raw_distinct": meta["raw_distinct"],
        "normalized_distinct": norm_codes,
        "tracker": meta["tracker_plans"],
        "dbf": meta["dbf_plans"],
        "crosswalk_codes": meta["crosswalk_codes"],
        "crosswalk_rows": meta["crosswalk_rows"],
        "reserve": meta["reserve_plans"],
        "draft": meta["draft_plans"],
        "tracker_only_count": len(t_only),
        "dbf_only_count": len(d_only),
        "tracker_only": t_only,
        "dbf_only": d_only,
        "scope_counts": dict(scope_counts),
        "status_counts": dict(status_counts),
        "in_scope_base": base_n,
        "in_scope_rider": rider_n,
        "missing_mapping_in_scope": missing_map,
        "missing_rate_source_in_scope": missing_src,
        "unresolved": unresolved,
        "relationship_candidates": len(rels),
        "plan_manifest_rows": len(pm),
        "git_exists": (CITIZENS / ".git").exists(),
        "qla_core_dir": (CITIZENS / "qla_core").exists(),
    }
    (GOV / "stage4a_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
