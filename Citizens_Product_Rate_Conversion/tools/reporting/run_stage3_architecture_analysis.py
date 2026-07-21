#!/usr/bin/env python3
"""
Stage 3 — Architecture baseline analysis (read-only; writes reports only into Citizens).
Never modifies CFIC_Rates. Never executes conversion. Never retargets imports.
"""

from __future__ import annotations

import ast
import csv
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

CITIZENS = Path(r"C:\Users\warren\Documents\GitHub\Warrenhughes1974\Citizens_Product_Rate_Conversion")
CFIC = Path(r"C:\Users\warren\Documents\GitHub\Warrenhughes1974\CFIC_Rates")

TECH_EXTS = {".py", ".cs", ".ps1", ".bat", ".sh", ".sql", ".yaml", ".yml", ".json", ".toml", ".ipynb", ".txt"}
QLA_IMPORT_RE = re.compile(r"(?:from\s+(qla_core(?:\.\w+)*)\s+import\s+([^\n]+)|import\s+(qla_core(?:\.\w+)*))")


def utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def write_csv(path: Path, columns: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def classify_py(rel: str) -> dict:
    """Disposition heuristics for Python assets."""
    r = rel.replace("\\", "/")
    if r.startswith("tools/inventory/") or r.startswith("tools/migration/"):
        return {
            "role": "Migration / inventory utility",
            "lifecycle": "active_tooling",
            "disposition": "RETAIN_ACTIVE",
            "safe": "SAFE_READ_ONLY",
            "historical": "N",
            "entry": "Y" if "run_stage" in r else "N",
        }
    if r.startswith("tools/reporting/") or "legacy__build" in r:
        return {
            "role": "Tracker / inventory builder",
            "lifecycle": "diagnostic",
            "disposition": "REFACTOR_LATER",
            "safe": "NOT_READY",
            "historical": "N",
            "entry": "Y",
        }
    if r.startswith("conversion/orchestration/"):
        if "legacy_cfic_paths" in r:
            return {
                "role": "Legacy path constants (pre-migration layout)",
                "lifecycle": "needs_refactor",
                "disposition": "REFACTOR_LATER",
                "safe": "DO_NOT_EXECUTE",
                "historical": "N",
                "entry": "N",
            }
        return {
            "role": "Legacy conversion orchestration",
            "lifecycle": "blocked_pending_engine",
            "disposition": "REFACTOR_LATER",
            "safe": "BLOCKED_BY_ENGINE",
            "historical": "N",
            "entry": "Y" if "package_cfic_rates" in r else "N",
        }
    if r.startswith("archive/"):
        return {
            "role": "Archived issue/dev script",
            "lifecycle": "historical",
            "disposition": "RETAIN_HISTORICAL",
            "safe": "HISTORICAL_ONLY",
            "historical": "Y",
            "entry": "N",
        }
    return {
        "role": "Unclassified Python",
        "lifecycle": "unknown",
        "disposition": "UNKNOWN",
        "safe": "DO_NOT_EXECUTE",
        "historical": "N",
        "entry": "N",
    }


def extract_imports(path: Path) -> tuple[list[str], list[str]]:
    local, external = [], []
    try:
        tree = ast.parse(path.read_text(encoding="utf-8", errors="replace"))
    except SyntaxError:
        return local, external
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                name = alias.name.split(".")[0]
                if name == "qla_core" or name.startswith("cfic_") or name in (
                    "cfic_paths", "cfic_inventory_core"
                ):
                    local.append(alias.name)
                else:
                    external.append(alias.name)
        elif isinstance(node, ast.ImportFrom):
            mod = node.module or ""
            if mod.startswith("qla_core") or mod.startswith("cfic_") or mod in (
                "cfic_paths", "cfic_inventory_core", "cfic_rate_publish",
                "cfic_reserve_build", "cfic_dbf_reader", "cfic_crosswalk",
            ):
                local.append(mod)
            else:
                external.append(mod)
    return sorted(set(local)), sorted(set(external))


def qla_details(path: Path) -> list[dict]:
    text = path.read_text(encoding="utf-8", errors="replace")
    findings = []
    for i, line in enumerate(text.splitlines(), 1):
        if "qla_core" not in line:
            continue
        m = QLA_IMPORT_RE.search(line)
        module = ""
        symbols = ""
        if m:
            module = m.group(1) or m.group(3) or "qla_core"
            symbols = (m.group(2) or "").strip()
        findings.append({
            "line": i,
            "line_text": line.strip()[:200],
            "module": module or "qla_core",
            "symbols": symbols,
        })
    return findings


def baseline_integrity() -> dict:
    inv = list(csv.DictReader((CITIZENS / "manifests/migration_inventory.csv").open(encoding="utf-8")))
    src_man = list(csv.DictReader((CITIZENS / "manifests/source_manifest.csv").open(encoding="utf-8")))
    missing, mismatch, ok = [], [], 0
    for r in src_man:
        d = CITIZENS.joinpath(*r["DESTINATION_RELATIVE_PATH"].split("/"))
        if not d.is_file():
            missing.append(r["DESTINATION_RELATIVE_PATH"])
            continue
        dig = sha256_file(d)
        if dig != r["DESTINATION_SHA256"]:
            mismatch.append(r["DESTINATION_RELATIVE_PATH"])
        else:
            ok += 1

    # Source snapshot compare
    post = json.loads((CITIZENS / "manifests/stage2b_post_source_snapshot.json").read_text(encoding="utf-8"))
    # Quick file count check on CFIC
    cfic_files = sum(1 for p in CFIC.rglob("*") if p.is_file()) if CFIC.is_dir() else -1

    approved_maps = list((CITIZENS / "mappings/approved").glob("*"))
    approved_maps = [p for p in approved_maps if p.is_file()]
    sensitive = list((CITIZENS / "quarantine/sensitive_review").glob("*"))
    cv_zips = list((CITIZENS / "source/original/cash_values").glob("*.zip"))
    extracted_cv = any(p.is_dir() and p.name.endswith("_CV") for p in (CITIZENS / "source/original/cash_values").iterdir()) if (CITIZENS / "source/original/cash_values").exists() else False

    result = {
        "timestamp_utc": utc_now(),
        "citizens_exists": CITIZENS.is_dir(),
        "inventory_rows": len(inv),
        "source_manifest_rows": len(src_man),
        "destinations_verified_ok": ok,
        "destinations_missing": len(missing),
        "destinations_hash_mismatch": len(mismatch),
        "mappings_approved_files": len(approved_maps),
        "sensitive_files": [p.name for p in sensitive if p.is_file()],
        "cash_value_zip_count": len(cv_zips),
        "cash_values_extracted": extracted_cv,
        "qla_core_in_citizens": (CITIZENS / "qla_core").exists(),
        "git_initialized": (CITIZENS / ".git").exists(),
        "nested_git_found": any(p.name == ".git" for p in CITIZENS.rglob(".git")),
        "cfic_file_count_now": cfic_files,
        "cfic_file_count_stage2b": post.get("file_count"),
        "cfic_total_bytes_stage2b": post.get("total_bytes"),
        "material_failure": False,
    }
    if len(inv) != 503 or len(src_man) != 380 or missing or mismatch or result["qla_core_in_citizens"] or result["git_initialized"]:
        result["material_failure"] = True
    if cfic_files != post.get("file_count"):
        result["cfic_file_count_delta"] = cfic_files - (post.get("file_count") or 0)
        # File count drift alone may be non-material (Excel lock); mark review not failure unless large
        if abs(result.get("cfic_file_count_delta", 0)) > 5:
            result["material_failure"] = True
    return result


def build_technical_manifest() -> list[dict]:
    rows = []
    aid = 0
    for p in sorted(CITIZENS.rglob("*")):
        if not p.is_file():
            continue
        ext = p.suffix.lower()
        name = p.name.lower()
        rel = p.relative_to(CITIZENS).as_posix()
        is_tech = (
            ext in TECH_EXTS
            or name.startswith("requirements")
            or name in ("pyproject.toml", "package.json", "setup.py", "pipfile")
        )
        # Skip huge data CSVs/JSON snapshots except config
        if ext in {".csv"} and not rel.startswith(("config/", "manifests/", "tools/", "conversion/", "reports/architecture")):
            if "requirements" not in name:
                continue
        if not is_tech and ext not in {".yaml", ".yml", ".json"}:
            continue
        if rel.startswith("staging/") or rel.startswith("output/") or rel.startswith("source/"):
            if ext not in {".py", ".yaml", ".yml"} and "requirements" not in name:
                continue
        # Skip OCR txt and SourceData txt as non-executable
        if ext == ".txt" and "requirements" not in name and not rel.startswith("tools/"):
            continue
        # Skip large snapshot JSON file_index noise - include only small config/report json
        if ext == ".json" and "file_index" not in rel:
            if p.stat().st_size > 500_000:
                continue

        aid += 1
        cls = classify_py(rel) if ext == ".py" else {
            "role": "Configuration / metadata" if ext in {".yaml", ".yml", ".json"} else "Requirements / package",
            "lifecycle": "active_config",
            "disposition": "RETAIN_ACTIVE",
            "safe": "SAFE_READ_ONLY",
            "historical": "Y" if rel.startswith("archive/") else "N",
            "entry": "N",
        }
        local_imp, ext_imp = ([], [])
        qla = "N"
        writes = "N"
        if ext == ".py":
            local_imp, ext_imp = extract_imports(p)
            qla = "Y" if any("qla_core" in x for x in local_imp) or "qla_core" in p.read_text(encoding="utf-8", errors="replace") else "N"
            text = p.read_text(encoding="utf-8", errors="replace")
            writes = "Y" if any(k in text for k in ("write_text", "open(", "to_csv", "shutil.copy", "mkdir", "Path.write")) else "N"
            # refine qla - only import counts as dependency for code
            qla = "Y" if any("qla_core" in x for x in local_imp) else "N"
            if "from qla_core" in text or "import qla_core" in text:
                qla = "Y"

        rows.append({
            "ASSET_ID": f"CIT-TECH-{aid:04d}",
            "CURRENT_RELATIVE_PATH": rel,
            "FILENAME": p.name,
            "LANGUAGE_OR_FORMAT": ext.lstrip(".") or "unknown",
            "PROBABLE_ROLE": cls["role"],
            "CURRENT_LIFECYCLE_STATUS": cls["lifecycle"],
            "PRIMARY_ENTRY_POINT_INDICATOR": cls["entry"],
            "IMPORTED_MODULES": ";".join(local_imp + ext_imp)[:500],
            "LOCAL_DEPENDENCIES": ";".join(local_imp)[:300],
            "EXTERNAL_DEPENDENCIES": ";".join(ext_imp)[:300],
            "QLA_CORE_DEPENDENCIES": qla,
            "INPUT_PATHS": "",
            "OUTPUT_PATHS": "",
            "CONFIGURATION_INPUTS": "",
            "SOURCE_DATA_DEPENDENCIES": "",
            "GENERATED_OUTPUT_DEPENDENCIES": "",
            "WRITES_FILES": writes,
            "MODIFIES_DATA": writes,
            "SAFE_TO_EXECUTE": cls["safe"],
            "CITIZENS_SPECIFIC": "Y",
            "APPEARS_REUSABLE": "Y" if "dbf_reader" in rel or "inventory" in rel else "N",
            "APPEARS_HISTORICAL": cls["historical"],
            "HAS_TESTS": "N",
            "HAS_DOCUMENTATION": "Y" if rel.startswith("tools/") else "N",
            "KNOWN_ISSUES": "Legacy CFIC_Rates paths; qla_core via sys.path" if rel.startswith("conversion/") else "",
            "RECOMMENDED_DISPOSITION": cls["disposition"],
        })
    return rows


def build_entry_points(tech_rows: list[dict]) -> list[dict]:
    entries = []
    catalog = [
        ("conversion/orchestration/package_cfic_rates.py", "Main reserve packaging orchestrator",
         "BLOCKED_BY_ENGINE", "Requires CFIC_Rates layout + qla_core"),
        ("conversion/orchestration/cfic_reserve_build.py", "Build Quik factor/key/member grids from reserve staging",
         "BLOCKED_BY_ENGINE", "Imports qla_core; hardcodes CFIC_Rates"),
        ("conversion/orchestration/cfic_rate_publish.py", "Publish Quik CSV load package",
         "BLOCKED_BY_ENGINE", "Imports qla_core.rate_dbf_writer"),
        ("conversion/orchestration/build_cfic_assumption_template.py", "Generate OBQ-2 assumption template from keys",
         "NOT_READY", "Hardcoded Output/Issue_Log paths"),
        ("conversion/orchestration/legacy_cfic_paths.py", "Path constants module",
         "DO_NOT_EXECUTE", "Not an entry point; obsolete layout"),
        ("tools/inventory/run_stage2a_inventory.py", "Stage 2A inventory (completed)",
         "SAFE_READ_ONLY", "Read-only vs CFIC_Rates"),
        ("tools/migration/run_stage2b_migration.py", "Stage 2B copy migration (completed)",
         "HISTORICAL_ONLY", "Do not re-run without authorization"),
        ("tools/reporting/_build_rate_load_tracker.py", "Rebuild rate-load tracker CSVs",
         "NOT_READY", "Parents[1] assumes old CFIC_Rates layout"),
        ("tools/inventory/legacy__build_plan_rate_inventory.py", "Build plan/rate PDF inventory",
         "NOT_READY", "Legacy path assumptions"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_03/scripts/extract_cfic_reserve_dbf.py",
         "Extract reserve DBF to staging", "HISTORICAL_ONLY", "Archived; paths point to CFIC_Rates"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_03/scripts/extract_cfic_plans_dbf.py",
         "Extract plans DBF", "HISTORICAL_ONLY", "Archived"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_03/scripts/validate_cfic_reserve_rates.py",
         "Validate reserve vs Access checkpoints", "HISTORICAL_ONLY", "Archived"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_03/scripts/emit_cfic_reserve_rates.py",
         "Thin wrapper to package_cfic_rates", "HISTORICAL_ONLY", "Archived"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_02/scripts/extract_cfic_pdf_rates.py",
         "PDF gross premium extract", "HISTORICAL_ONLY", "OCR/PDF pilot"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_02/scripts/emit_cfic_pdf_rates.py",
         "Emit QuikGps from PDF staging", "BLOCKED_BY_ENGINE", "qla_core + archived"),
        ("archive/legacy_cfic_rates/issues/CFIC_Issue_01/scripts/extract_cfic_green_sheets.py",
         "OCR green-sheet extract", "DO_NOT_EXECUTE", "OCR FAIL; not authoritative"),
    ]
    for path, purpose, verdict, notes in catalog:
        p = CITIZENS / path
        qla = "N"
        cfic_ref = "N"
        qla_mig = "N"
        if p.is_file() and p.suffix == ".py":
            text = p.read_text(encoding="utf-8", errors="replace")
            qla = "Y" if "qla_core" in text and ("import" in text or "from qla_core" in text) else ("Y" if "from qla_core" in text or "import qla_core" in text else "N")
            if re.search(r"from qla_core|import qla_core", text):
                qla = "Y"
            else:
                qla = "N"
            cfic_ref = "Y" if "CFIC_Rates" in text else "N"
            qla_mig = "Y" if "QLA_Migration" in text else "N"
        entries.append({
            "ENTRY_POINT_PATH": path,
            "INTENDED_PURPOSE": purpose,
            "REQUIRED_WORKING_DIRECTORY": "Citizens project root (intended future); currently assumes CFIC_Rates",
            "REQUIRED_COMMAND": f"python {path}",
            "REQUIRED_ARGUMENTS": "See script --help when runnable",
            "REQUIRED_ENVIRONMENT_VARIABLES": "None documented",
            "REQUIRED_PACKAGES": "pandas/openpyxl/easyocr/pymupdf as applicable; qla_core external",
            "REQUIRED_SOURCE_FILES": "DBF/crosswalk/staging depending on script",
            "EXPECTED_OUTPUTS": "Staging CSVs / Quik CSVs / trackers / reports",
            "OUTPUT_STATUS": "draft_or_historical",
            "CURRENTLY_RUNNABLE": "N",
            "WRITES_OUTSIDE_CITIZENS": "UNKNOWN_IF_RUN" if "conversion/" in path else "N",
            "REFERENCES_CFIC_RATES": cfic_ref,
            "REFERENCES_QLA_MIGRATION": qla_mig,
            "REFERENCES_QLA_CORE": qla,
            "UNDOCUMENTED_ASSUMPTIONS": "Y",
            "SAFE_TO_RUN_VERDICT": verdict,
            "NOTES": notes,
            "EXISTS": "Y" if p.is_file() else "N",
        })
    return entries


def build_qla_matrix() -> list[dict]:
    rows = []
    for p in sorted(CITIZENS.rglob("*.py")):
        text = p.read_text(encoding="utf-8", errors="replace")
        if not re.search(r"from qla_core|import qla_core", text):
            continue
        rel = p.relative_to(CITIZENS).as_posix()
        for f in qla_details(p):
            if not re.search(r"from qla_core|import qla_core", f["line_text"]):
                continue
            hist = "Y" if rel.startswith("archive/") else "N"
            classification = "HISTORICAL_ONLY" if hist == "Y" else "REQUIRED_ENGINE_API"
            rows.append({
                "CITIZENS_FILE": rel,
                "LINE_NUMBER": f["line"],
                "IMPORTED_QLA_CORE_MODULE": f["module"],
                "IMPORTED_CLASS_OR_FUNCTION": f["symbols"] or "(module import)",
                "HOW_USED": "Schema/factor/key/member/writer for Quik rate emit",
                "REQUIRED_INPUTS": "Normalized reserve/premium staging + crosswalk assumptions",
                "PRODUCED_OUTPUTS": "Quik factor/key/member CSV structures",
                "INTERFACE_GENERIC": "Y",
                "CITIZENS_SPECIFIC_ASSUMPTIONS_PASSED": "Y",
                "APPEARS_STABLE": "UNKNOWN",
                "LOCAL_ALTERNATIVE": "N",
                "RISK_IF_INTERFACE_CHANGES": "HIGH — blocks all Quik emit",
                "PROPOSED_INTEGRATION_METHOD": "Pinned Enterprise Engine package (Option A)",
                "DEPENDENCY_CLASSIFICATION": classification,
            })
    return rows


def build_path_report() -> list[dict]:
    patterns = [
        (r"CFIC_Rates", "legacy_project_name"),
        (r"Citizens_Product_Rate_Conversion", "citizens_project_name"),
        (r"QLA_Migration", "warren_project"),
        (r"Warrenhughes1974", "repo_name"),
        (r"C:\\\\Users", "absolute_path"),
        (r"C:/Users", "absolute_path"),
        (r"\bOutput\b", "output_folder"),
        (r"\boutput\b", "output_folder"),
        (r"QLA_Migration", "warren_project"),
        (r"Issue_Log", "legacy_issue_folder"),
        (r"extracted_reserve", "legacy_staging"),
        (r"extracted_plans", "legacy_staging"),
        (r"extracted_pdf_rates", "legacy_staging"),
        (r"CFIC_Cash_Values", "legacy_source_folder"),
        (r"qla_core", "engine_module"),
        (r"\bCSO\b", "cso_reference"),
        (r"sys\.path", "sys_path_manipulation"),
        (r"parents\[\d+\]", "parent_traversal"),
    ]
    rows = []
    for p in sorted(CITIZENS.rglob("*.py")):
        rel = p.relative_to(CITIZENS).as_posix()
        lines = p.read_text(encoding="utf-8", errors="replace").splitlines()
        for i, line in enumerate(lines, 1):
            for pat, rtype in patterns:
                if re.search(pat, line):
                    disposition = "MOVE_TO_CONFIG"
                    if rtype == "cso_reference":
                        disposition = "HISTORICAL_NO_CHANGE"
                    if rtype == "engine_module":
                        disposition = "REPLACE_WITH_ENGINE_API"
                    if rel.startswith("archive/") or rel.startswith("tools/migration") or rel.startswith("tools/inventory"):
                        disposition = "HISTORICAL_NO_CHANGE" if "Stage" in rel or "archive" in rel else disposition
                    rows.append({
                        "FILE": rel,
                        "LINE_NUMBER": i,
                        "REFERENCE": re.search(pat, line).group(0)[:80],
                        "REFERENCE_TYPE": rtype,
                        "CURRENT_PURPOSE": line.strip()[:160],
                        "REQUIRED_FUTURE_DISPOSITION": disposition,
                        "PROPOSED_CONFIGURATION_KEY": {
                            "legacy_project_name": "paths.legacy_cfic_rates_readonly",
                            "absolute_path": "paths.project_root",
                            "output_folder": "paths.output_rates",
                            "legacy_staging": "paths.staging",
                            "engine_module": "engine.package",
                            "sys_path_manipulation": "engine.install_method",
                            "parent_traversal": "paths.project_root",
                            "legacy_issue_folder": "paths.issues",
                            "legacy_source_folder": "paths.source.cash_values",
                            "warren_project": "REMOVE — out of scope",
                            "cso_reference": "N/A — documentation only",
                            "citizens_project_name": "KEEP_RELATIVE",
                            "repo_name": "paths.monorepo_root_optional",
                        }.get(rtype, "REVIEW"),
                        "DEVELOPMENT_ISSUE_REQUIRED": "CIT-ARCH-001" if disposition == "MOVE_TO_CONFIG" else "CIT-ENGINE-001" if disposition == "REPLACE_WITH_ENGINE_API" else "",
                        "RISK": "HIGH" if rtype in ("absolute_path", "sys_path_manipulation", "engine_module", "legacy_project_name") else "MEDIUM",
                    })
    return rows


def plan_universe() -> list[dict]:
    tracker_path = CITIZENS / "discovery/rates/CFIC_Rate_Load_Tracker.csv"
    plans_path = CITIZENS / "staging/normalized_plans/staging/plans_master.csv"
    # find tracker
    if not tracker_path.exists():
        candidates = list((CITIZENS / "discovery").rglob("*Tracker*.csv"))
        tracker_path = candidates[0] if candidates else tracker_path
    if not plans_path.exists():
        candidates = list((CITIZENS / "staging").rglob("plans_master.csv"))
        plans_path = candidates[0] if candidates else plans_path

    tracker_plans = set()
    tracker_meta = {}
    if tracker_path.exists():
        for r in csv.DictReader(tracker_path.open(encoding="utf-8", errors="replace")):
            code = (r.get("cfic_plan_code") or r.get("CFIC_PLAN") or "").strip()
            if code:
                tracker_plans.add(code)
                tracker_meta[code] = r

    dbf_plans = set()
    dbf_meta = {}
    if plans_path.exists():
        for r in csv.DictReader(plans_path.open(encoding="utf-8", errors="replace")):
            # try common columns
            code = (r.get("PLAN") or r.get("plan") or r.get("CFIC_PLAN") or r.get("plan_code") or "").strip()
            if not code:
                # first column
                code = list(r.values())[0].strip() if r else ""
            if code:
                dbf_plans.add(code)
                dbf_meta[code] = r

    reserve_dirs = set()
    reserve_root = CITIZENS / "staging/normalized_rates/reserve/staging"
    if reserve_root.exists():
        reserve_dirs = {d.name for d in reserve_root.iterdir() if d.is_dir()}

    draft_plans = set()
    # emit summary may list plans
    emit = CITIZENS / "reports/audit/emit_summary.json"
    if emit.exists():
        data = json.loads(emit.read_text(encoding="utf-8"))
        for p in data.get("plans", []):
            draft_plans.add(str(p).strip())

    # crosswalk - try openpyxl or skip if unavailable
    crosswalk_plans = set()
    xlsx = CITIZENS / "mappings/working/plans/Citizens_Plan_Crosswalk.xlsx"
    if not xlsx.exists():
        xlsx = next((CITIZENS / "mappings/working").rglob("*Crosswalk*.xlsx"), None) or next((CITIZENS / "mappings/working").rglob("*Crosswak*.xlsx"), None)
    if xlsx and Path(xlsx).exists():
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            # find plan column
            for row in ws.iter_rows(min_row=2, values_only=True):
                for val in row:
                    if val is None:
                        continue
                    s = str(val).strip()
                    if not s:
                        continue
                    # split grouped codes
                    for part in re.split(r"[,;/]", s):
                        part = part.strip()
                        if 1 <= len(part) <= 8 and re.match(r"^[A-Za-z0-9.\-]+$", part):
                            # heuristic: only if looks like plan code column content
                            pass
            # Better: use first column typically CFIC plan
            ws2 = wb.active
            rows_x = list(ws2.iter_rows(min_row=2, values_only=True))
            for row in rows_x:
                if not row:
                    continue
                cell = row[0]
                if cell is None:
                    continue
                for part in re.split(r"[,;/]", str(cell)):
                    part = part.strip()
                    if part and re.match(r"^[A-Za-z0-9.\-]{1,8}$", part):
                        crosswalk_plans.add(part)
            wb.close()
        except Exception as e:
            crosswalk_plans = set()  # leave empty with note later

    all_codes = sorted(tracker_plans | dbf_plans | crosswalk_plans | reserve_dirs | draft_plans)
    out = []
    for code in all_codes:
        in_t = code in tracker_plans
        in_d = code in dbf_plans
        in_c = code in crosswalk_plans
        in_r = code in reserve_dirs
        in_o = code in draft_plans
        flags = [in_t, in_d, in_c, in_r, in_o]
        if all([in_t, in_d]) or sum(flags) >= 3:
            status = "MATCHED"
        elif in_t and not in_d:
            status = "TRACKER_ONLY"
        elif in_d and not in_t:
            status = "DBF_ONLY"
        elif in_c and not in_t and not in_d:
            status = "CROSSWALK_ONLY"
        elif in_r and not in_t:
            status = "SOURCE_ONLY"
        else:
            status = "REQUIRES_INTERNAL_REVIEW"
        meta = tracker_meta.get(code, {})
        out.append({
            "SOURCE_PLAN_CODE": code,
            "IN_TRACKER": "Y" if in_t else "N",
            "IN_PLAN_DBF": "Y" if in_d else "N",
            "IN_CROSSWALK": "Y" if in_c else "N",
            "IN_RATE_REQUIREMENTS": "Y" if in_t else "N",  # tracker derived from requirements
            "IN_RESERVE_STAGING": "Y" if in_r else "N",
            "IN_DRAFT_OUTPUT": "Y" if in_o else "N",
            "PLAN_NAME": meta.get("product_family", ""),
            "PRODUCT_FAMILY": meta.get("product_family", ""),
            "POSSIBLE_ALIAS": "",
            "SOURCE_AUTHORITY_STATUS": "UNKNOWN",
            "RECONCILIATION_STATUS": status,
            "NOTES": "",
        })
    return out


def rate_universe() -> list[dict]:
    cats = [
        ("GROSS_PREMIUM", "Access + PDF pilot", "PENDING_REVIEW", "Y", "Y", "N", "N", "Y", "N"),
        ("CASH_VALUE", "cifi0007.DBF RL_CASHVAL", "PENDING_REVIEW", "Y", "Y", "Y", "Y", "Y", "Y"),
        ("NET_PREMIUM", "cifi0007.DBF RL_NETPREM", "PENDING_REVIEW", "Y", "Y", "Y", "Y", "Y", "Y"),
        ("TERM_RESERVE", "cifi0007.DBF terminal reserve", "PENDING_REVIEW", "Y", "Y", "Y", "Y", "Y", "Y"),
        ("MEAN_RESERVE", "cifi0007.DBF mean fields", "UNKNOWN", "Y", "N", "N", "N", "N", "N"),
        ("PAID_UP", "cifi0007.DBF PUP", "PENDING_REVIEW", "Y", "Y", "Y", "Y", "Y", "Y"),
        ("EXTENDED_TERM", "green sheets / OBQ", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("DIVIDEND", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("DIVIDEND_INTEREST", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("LOAN_INTEREST", "cifi0004.dbf IR1-IR8", "PENDING_REVIEW", "Y", "N", "N", "N", "N", "N"),
        ("CREDITED_INTEREST", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("GUARANTEED_INTEREST", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("CURRENT_INTEREST", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("COST_OF_INSURANCE", "requirements catalog", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("EXPENSE_CHARGE", "requirements catalog", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("POLICY_FEE", "cifi0004.dbf", "PENDING_REVIEW", "Y", "N", "N", "N", "N", "N"),
        ("PREMIUM_LOAD", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("SURRENDER_CHARGE", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("MODAL_FACTOR", "requirements catalog", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("RIDER_PREMIUM", "Access rider CSVs", "UNKNOWN", "Y", "N", "N", "N", "N", "N"),
        ("GUIDELINE_PREMIUM", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("SEVEN_PAY", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("TARGET_PREMIUM", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("SETTLEMENT_FACTOR", "", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
        ("OTHER", "requirements catalog Other", "UNKNOWN", "N", "N", "N", "N", "N", "N"),
    ]
    rows = []
    for code, src, auth, extr, norm, map_, conv, val, draft in cats:
        rows.append({
            "RATE_TYPE_CODE": code,
            "SOURCE_IDENTIFIED": src or "N",
            "SOURCE_AUTHORITY_STATUS": auth,
            "EXISTING_EXTRACTION": extr,
            "EXISTING_NORMALIZATION": norm,
            "EXISTING_MAPPING": map_,
            "EXISTING_CONVERSION": conv,
            "EXISTING_VALIDATION": val,
            "EXISTING_DRAFT_OUTPUT": draft,
            "KNOWN_MISSING_DATA": "Y" if auth == "UNKNOWN" else "Partial",
            "KNOWN_BLOCKERS": "OBQ-1/OBQ-2; engine pin; path retarget" if conv == "Y" else "Source authority undecided",
            "RECOMMENDED_NEXT_ISSUE": "CIT-RATE-001" if code in ("CASH_VALUE", "TERM_RESERVE", "PAID_UP", "NET_PREMIUM") else "CIT-DATA-001",
        })
    return rows


def data_assets() -> list[dict]:
    sets = [
        ("source/original/dbf/", "AUTHORITATIVE_SOURCE", "PENDING_REVIEW", "FoxPro DBFs"),
        ("source/original/cash_values/", "AUTHORITATIVE_SOURCE", "PENDING_REVIEW", "CV ZIP archives unextracted"),
        ("source/original/access/", "AUTHORITATIVE_SOURCE", "PENDING_REVIEW", "Access MDB/ZIP"),
        ("source/extracts/access/", "CONTROLLED_EXTRACT", "PENDING_REVIEW", "Access CSV exports"),
        ("source/product_documents/", "AUTHORITATIVE_SOURCE", "PENDING_REVIEW", "Rate sheet PDFs"),
        ("mappings/working/", "WORKING_MAPPING", "working", "Crosswalk + catalog + assumptions"),
        ("mappings/approved/", "WORKING_MAPPING", "empty", "Empty — do not treat as approved"),
        ("staging/normalized_rates/reserve/", "NORMALIZED_STAGING", "working", "Reserve grids"),
        ("staging/normalized_plans/", "NORMALIZED_STAGING", "working", "Plans master extract"),
        ("output/csv/draft_pre_migration/", "DRAFT_OUTPUT", "historical", "Draft Quik* — not production"),
        ("validation/", "VALIDATION_EVIDENCE", "working", "Parity and issue evidence"),
        ("reports/audit/", "VALIDATION_EVIDENCE", "historical", "Emit summaries"),
        ("archive/legacy_cfic_rates/SourceData_11-18-2024/", "HISTORICAL_OUTPUT", "HISTORICAL_PENDING_REVIEW", "Legacy dump"),
        ("archive/legacy_cfic_rates/ocr_extract/", "GENERATED_DISPOSABLE", "historical", "OCR text — not authority"),
        ("archive/legacy_cfic_rates/green_sheet_pilot/", "HISTORICAL_OUTPUT", "historical", "OCR pilot FAIL"),
        ("quarantine/sensitive_review/", "UNKNOWN", "quarantined", "PII / scope review"),
        ("quarantine/duplicate_review/", "HISTORICAL_OUTPUT", "quarantined", "Duplicates for audit"),
        ("discovery/rates/", "WORKING_MAPPING", "working", "Rate load trackers"),
    ]
    return [{
        "DATA_ASSET_SET": path,
        "CLASSIFICATION": cls,
        "APPROVAL_STATUS": appr,
        "NOTES": notes,
    } for path, cls, appr, notes in sets]


def unsafe_scripts() -> list[dict]:
    return [
        {
            "SCRIPT_PATH": "conversion/orchestration/package_cfic_rates.py",
            "UNSAFE_BEHAVIORS": "Assumes CFIC_Rates layout; sys.path to monorepo; writes Output/rates; mixes extract/validate/publish",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "Refactor under CIT-ARCH-001 + CIT-ENGINE-001 before any run",
        },
        {
            "SCRIPT_PATH": "conversion/orchestration/cfic_reserve_build.py",
            "UNSAFE_BEHAVIORS": "Hardcodes REPO_ROOT/CFIC_Rates; imports qla_core via sys.path",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "Engine pin + path config",
        },
        {
            "SCRIPT_PATH": "conversion/orchestration/cfic_rate_publish.py",
            "UNSAFE_BEHAVIORS": "Publishes Quik CSVs without Stage gates; qla_core dependency",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "Gate behind validation + run_manifest",
        },
        {
            "SCRIPT_PATH": "conversion/orchestration/legacy_cfic_paths.py",
            "UNSAFE_BEHAVIORS": "Points to obsolete CFIC folder names relative to script parent",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "Replace with config/source_locations.yaml",
        },
        {
            "SCRIPT_PATH": "conversion/orchestration/build_cfic_assumption_template.py",
            "UNSAFE_BEHAVIORS": "Hardcoded Output and Issue_Log paths",
            "SEVERITY": "MEDIUM",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "Retarget after path config",
        },
        {
            "SCRIPT_PATH": "archive/legacy_cfic_rates/issues/CFIC_Issue_01/scripts/extract_cfic_green_sheets.py",
            "UNSAFE_BEHAVIORS": "Uses OCR as rate source; pilot FAIL; not authoritative when DBF exists",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "DO_NOT_EXECUTE as source authority",
        },
        {
            "SCRIPT_PATH": "archive/legacy_cfic_rates/issues/CFIC_Issue_02/scripts/emit_cfic_pdf_rates.py",
            "UNSAFE_BEHAVIORS": "Emit without approved mapping; qla_core; archived paths",
            "SEVERITY": "HIGH",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "HISTORICAL_ONLY until gross-premium issue gated",
        },
        {
            "SCRIPT_PATH": "tools/reporting/_build_rate_load_tracker.py",
            "UNSAFE_BEHAVIORS": "parents[1] path assumes old tree; may write discovery trackers incorrectly",
            "SEVERITY": "MEDIUM",
            "SAFE_TO_RUN_NOW": "N",
            "RECOMMENDED_ACTION": "REFACTOR_LATER",
        },
    ]


def main() -> int:
    print("Stage 3 baseline integrity...")
    base = baseline_integrity()
    arch = CITIZENS / "reports/architecture"
    arch.mkdir(parents=True, exist_ok=True)
    docs = CITIZENS / "docs/architecture"
    docs.mkdir(parents=True, exist_ok=True)

    integrity_md = f"""# Stage 3 Baseline Integrity Report

**Generated:** {base['timestamp_utc']}

## Result: {'FAIL' if base['material_failure'] else 'PASS'}

| Check | Value |
|-------|------:|
| Inventory rows | {base['inventory_rows']} (expect 503) |
| Source manifest rows | {base['source_manifest_rows']} (expect 380) |
| Destinations hash OK | {base['destinations_verified_ok']} |
| Destinations missing | {base['destinations_missing']} |
| Destinations hash mismatch | {base['destinations_hash_mismatch']} |
| mappings/approved files | {base['mappings_approved_files']} |
| Sensitive files | {', '.join(base['sensitive_files'])} |
| Cash-value ZIPs | {base['cash_value_zip_count']} |
| Cash values extracted | {base['cash_values_extracted']} |
| qla_core inside Citizens | {base['qla_core_in_citizens']} |
| Git initialized | {base['git_initialized']} |
| Nested .git | {base['nested_git_found']} |
| CFIC_Rates file count now | {base['cfic_file_count_now']} |
| CFIC_Rates file count Stage 2B | {base['cfic_file_count_stage2b']} |

## Material Failure

{base['material_failure']}

If FAIL, stop Stage 3 architecture work until baseline is restored.
"""
    (arch / "Stage3_Baseline_Integrity_Report.md").write_text(integrity_md, encoding="utf-8")
    if base["material_failure"]:
        print("MATERIAL FAILURE — stopping")
        return 2

    print("Technical asset manifest...")
    tech = build_technical_manifest()
    write_csv(CITIZENS / "manifests/technical_asset_manifest.csv", list(tech[0].keys()) if tech else ["ASSET_ID"], tech)

    print("Entry points...")
    entries = build_entry_points(tech)
    write_csv(arch / "entry_point_inventory.csv", list(entries[0].keys()), entries)

    print("qla_core matrix...")
    qla = build_qla_matrix()
    write_csv(arch / "qla_core_dependency_matrix.csv", list(qla[0].keys()) if qla else ["CITIZENS_FILE"], qla)

    print("Path report...")
    paths = build_path_report()
    write_csv(arch / "legacy_path_and_reference_report.csv", list(paths[0].keys()) if paths else ["FILE"], paths)

    print("Data assets...")
    das = data_assets()
    write_csv(CITIZENS / "manifests/data_asset_classification.csv", list(das[0].keys()), das)

    print("Plan universe...")
    plans = plan_universe()
    write_csv(arch / "plan_universe_reconciliation.csv", list(plans[0].keys()) if plans else ["SOURCE_PLAN_CODE"], plans)

    print("Rate universe...")
    rates = rate_universe()
    write_csv(arch / "rate_universe_baseline.csv", list(rates[0].keys()), rates)

    print("Unsafe scripts...")
    unsafe = unsafe_scripts()
    write_csv(arch / "unsafe_script_register.csv", list(unsafe[0].keys()), unsafe)

    # Persist summary stats for final report
    summary = {
        "timestamp_utc": utc_now(),
        "baseline_pass": not base["material_failure"],
        "tech_assets": len(tech),
        "tech_active": sum(1 for t in tech if t["RECOMMENDED_DISPOSITION"] in ("RETAIN_ACTIVE", "REFACTOR_LATER", "MOVE_TO_TOOLING")),
        "tech_historical": sum(1 for t in tech if t["APPEARS_HISTORICAL"] == "Y" or t["RECOMMENDED_DISPOSITION"] == "RETAIN_HISTORICAL"),
        "unsafe_scripts": len(unsafe),
        "qla_import_rows": len(qla),
        "qla_files": len(set(r["CITIZENS_FILE"] for r in qla)),
        "path_findings": len(paths),
        "plan_universe_rows": len(plans),
        "plan_tracker": sum(1 for p in plans if p["IN_TRACKER"] == "Y"),
        "plan_dbf": sum(1 for p in plans if p["IN_PLAN_DBF"] == "Y"),
        "plan_crosswalk": sum(1 for p in plans if p["IN_CROSSWALK"] == "Y"),
        "plan_reserve": sum(1 for p in plans if p["IN_RESERVE_STAGING"] == "Y"),
        "entry_points": len(entries),
        "entry_blocked": sum(1 for e in entries if e["SAFE_TO_RUN_VERDICT"] in ("BLOCKED_BY_ENGINE", "DO_NOT_EXECUTE", "NOT_READY", "HISTORICAL_ONLY")),
    }
    (arch / "stage3_analysis_summary.json").write_text(json.dumps(summary, indent=2), encoding="utf-8")
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
