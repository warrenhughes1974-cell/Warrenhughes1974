"""Build QUIKAINT.DBF (QLA interest rate table) from PFSA rate workbooks.

Sources (per business direction, 2026-06-12):
  - History 2016-2025: "PFSA Interest Rate Crediting History - Copy.xlsx"
      sheet "PFSA Rates" with two blocks:
        * "First Year/Promo Rates"     -> MINTRATE1 (current products only)
        * "Renewal Crediting Rates"    -> MINTRATE  (current + legacy plans)
      Rates stored as decimal fractions (0.05 == 5%).
  - 2026 only: "PFSA_Interest_Rates.xlsx" sheet "Current Products"
      (2026 Renewal Crediting -> MINTRATE, 2026 Promotional -> MINTRATE1).
  - Plan crosswalk (PFSA plan number -> QLA plan codes) comes from the
    "QLAdmin Plan" column of PFSA_Interest_Rates.xlsx (both blocks).

Rules (unchanged from 2026-06-09 build):
  - MPLAN     <- QLAdmin plan code (slash-delimited lists expand per code)
  - MINTRATE  <- Crediting / Renewal rate for the year
  - MINTRATE1 <- Promotional rate; falls back to crediting when absent
  - MEFFDATE  <- Jan 1 of the rate year; the first record per plan is
                 back-dated to 19000101
  - A record is emitted only when the (MINTRATE, MINTRATE1) pair changes vs
    the plan's previously emitted record (years may be skipped).
  - "No match provided" plans and non-numeric rates (N/A, minimum) skip.

The DBF header is cloned from the example QUIKAINT.DBF so the physical layout
(version byte, header length 162, record length 29) matches QLA exactly.
"""

import csv
import datetime
import os
import struct
import sys

import openpyxl

HISTORY_XLSX = r"c:\Users\warren\Desktop\PFSA Interest Rate Crediting History - Copy.xlsx"
RATES_2026_XLSX = r"c:\Users\warren\Desktop\PFSA_Interest_Rates.xlsx"
EXAMPLE_DBF = r"q:\PFSA\Jessica_PFSA_FinalMergedDataTesting\QUIKAINT.DBF"
OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_DBF = os.path.join(OUT_DIR, "QUIKAINT.DBF")
OUT_TRACE = os.path.join(OUT_DIR, "quikaint_emit_trace.csv")
OUT_APPEND = os.path.join(OUT_DIR, "quikaint_append.csv")

HEADER_LEN = 162
RECORD_LEN = 29

# Extra QLAdmin plan codes not present in the spreadsheet but confirmed by
# business to share a PFSA plan's rate set (keyed by PFSA Plan number).
#   2026-06-09: A107NA added to PFSA plan 107 (same rates as A103RO / A108SP).
EXTRA_PLAN_CODES = {
    107: ["A107NA"],
}


def parse_rate(value):
    """Return float percent or None for blank / N-A / non-numeric text."""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.upper() in ("N/A", "NA"):
        return None
    if s.endswith("%"):
        s = s[:-1]
    try:
        v = float(s)
    except ValueError:
        return None  # e.g. 'minimum' notation in legacy 2016 column
    # Excel sometimes stores percents as fractions (0.03 == 3%)
    return v * 100.0 if v < 1.0 else v


def split_plans(cell):
    if cell is None:
        return []
    s = str(cell).strip()
    if not s or "no match" in s.lower():
        return []
    plans = [p.strip() for p in s.split("/") if p.strip()]
    for p in plans:
        if len(p) > 6:
            raise ValueError(f"Plan code longer than C(6): {p!r}")
    return plans


def read_crosswalk():
    """PFSA plan number -> (plan_name, [qla_codes]) from PFSA_Interest_Rates.xlsx."""
    wb = openpyxl.load_workbook(RATES_2026_XLSX, data_only=True)
    ws = wb["Current Products"]
    rows = list(ws.iter_rows(values_only=True))

    crosswalk = {}

    # Current products block: header sheet row 2, data rows 3..
    for r in rows[2:]:
        if r[0] is None:
            break
        plans = split_plans(r[1])
        if not plans:
            continue
        plans += [c for c in EXTRA_PLAN_CODES.get(r[0], []) if c not in plans]
        crosswalk[int(r[0])] = (str(r[2]), plans)

    # Legacy block: locate its header row ("Group" in column D)
    legacy_start = None
    for i, r in enumerate(rows):
        if r[3] == "Group":
            legacy_start = i + 1
            break
    if legacy_start is None:
        raise RuntimeError("Legacy block header not found")

    for r in rows[legacy_start:]:
        if r[0] is None:
            continue
        plans = split_plans(r[1])
        if not plans:
            continue
        plans += [c for c in EXTRA_PLAN_CODES.get(r[0], []) if c not in plans]
        crosswalk[int(r[0])] = (str(r[2]), plans)

    return crosswalk


def read_2026_rates():
    """PFSA plan number -> (renewal, promo) for 2026 from PFSA_Interest_Rates.xlsx."""
    wb = openpyxl.load_workbook(RATES_2026_XLSX, data_only=True)
    ws = wb["Current Products"]
    rows = list(ws.iter_rows(values_only=True))

    rates = {}

    # Current block: N = 2026 Renewal Crediting (idx 13), P = 2026 Promotional (idx 15)
    for r in rows[2:]:
        if r[0] is None:
            break
        renewal = parse_rate(r[13])
        promo = parse_rate(r[15])
        if renewal is not None:
            rates[int(r[0])] = (renewal, promo)

    # Legacy block: 2026 column (idx 12) — currently all N/A, handled generically
    legacy_start = None
    for i, r in enumerate(rows):
        if r[3] == "Group":
            legacy_start = i + 1
            break
    for r in rows[legacy_start:]:
        if r[0] is None:
            continue
        renewal = parse_rate(r[12])
        if renewal is not None:
            rates[int(r[0])] = (renewal, None)

    return rates


def _read_history_block(rows, title):
    """Return ({pfsa: {year: rate}}, header_years) for the block under `title`."""
    block_start = None
    for i, r in enumerate(rows):
        if r[0] is not None and str(r[0]).strip() == title:
            block_start = i + 1
            break
    if block_start is None:
        raise RuntimeError(f"History block not found: {title!r}")

    header = rows[block_start]
    if str(header[0]).strip() != "Plan Code":
        raise RuntimeError(f"Unexpected header under {title!r}: {header[0]!r}")
    year_cols = {}
    for idx, cell in enumerate(header):
        if isinstance(cell, (int, float)) and 1990 < int(cell) < 2100:
            year_cols[idx] = int(cell)

    data = {}
    for r in rows[block_start + 1:]:
        if r[0] is None:
            break
        try:
            pfsa = int(r[0])
        except (TypeError, ValueError):
            break
        per_year = {}
        for idx, year in year_cols.items():
            rate = parse_rate(r[idx])
            if rate is not None:
                per_year[year] = rate
        data[pfsa] = per_year
    return data


def read_history():
    """PFSA plan number -> {year: (renewal, promo_or_None)} from the history workbook."""
    wb = openpyxl.load_workbook(HISTORY_XLSX, data_only=True)
    ws = wb["PFSA Rates"]
    rows = list(ws.iter_rows(values_only=True))

    promo = _read_history_block(rows, "First Year/Promo Rates")
    renewal = _read_history_block(rows, "Renewal Crediting Rates")

    history = {}
    for pfsa, per_year in renewal.items():
        promo_years = promo.get(pfsa, {})
        history[pfsa] = {
            year: (rate, promo_years.get(year))
            for year, rate in sorted(per_year.items())
        }
    return history


def build_entries():
    """Return list of (pfsa_plan, plan_name, [qla_codes], [(year, crediting, promo)])."""
    crosswalk = read_crosswalk()
    history = read_history()
    rates_2026 = read_2026_rates()

    entries = []
    for pfsa, per_year in sorted(history.items()):
        if pfsa not in crosswalk:
            continue  # "No match provided" plans (12, 14, 83, 84, 99)
        name, plans = crosswalk[pfsa]
        series = [(year, renewal, promo) for year, (renewal, promo) in sorted(per_year.items())]
        if pfsa in rates_2026:
            renewal_26, promo_26 = rates_2026[pfsa]
            series.append((2026, renewal_26, promo_26))
        entries.append((pfsa, name, plans, series))
    return entries


def build_records(entries):
    """Apply change-only emission. Returns list of dicts."""
    records = []
    for pfsa, name, plans, series in entries:
        emitted_pair = None
        plan_records = []
        for year, crediting, promo in series:
            if crediting is None:
                continue  # N/A year - skip
            mintrate1 = promo if promo is not None else crediting
            pair = (round(crediting, 4), round(mintrate1, 4))
            if pair == emitted_pair:
                continue
            emitted_pair = pair
            plan_records.append((year, pair[0], pair[1]))
        for code in plans:
            for idx, (year, mintrate, mintrate1) in enumerate(plan_records):
                # First record per plan is effective from 01/01/1900 through
                # to its first rate change.
                meffdate = "19000101" if idx == 0 else f"{year}0101"
                records.append({
                    "PFSA_PLAN": pfsa,
                    "PLAN_NAME": name,
                    "MPLAN": code,
                    "MEFFDATE": meffdate,
                    "MINTRATE": mintrate,
                    "MINTRATE1": mintrate1,
                })
    records.sort(key=lambda r: (r["MPLAN"], r["MEFFDATE"]))
    return records


def _read_header_template():
    """Clone the 162-byte header from the example DBF; fall back to the
    previously generated local DBF when the network drive is unavailable."""
    for path in (EXAMPLE_DBF, OUT_DBF):
        try:
            with open(path, "rb") as f:
                header = bytearray(f.read(HEADER_LEN))
            if len(header) == HEADER_LEN:
                return header
        except OSError:
            continue
    raise RuntimeError(
        "No DBF header template available: example DBF unreachable and no prior local build"
    )


def write_dbf(records):
    header = _read_header_template()

    today = datetime.date.today()
    header[1] = today.year - 1900
    header[2] = today.month
    header[3] = today.day
    struct.pack_into("<L", header, 4, len(records))

    with open(OUT_DBF, "wb") as f:
        f.write(header)
        for r in records:
            rec = (
                b" "
                + r["MPLAN"].ljust(6).encode("ascii")
                + r["MEFFDATE"].encode("ascii")
                + f"{r['MINTRATE']:7.4f}".encode("ascii")
                + f"{r['MINTRATE1']:7.4f}".encode("ascii")
            )
            assert len(rec) == RECORD_LEN, (len(rec), r)
            f.write(rec)
        f.write(b"\x1a")


def write_trace(records):
    with open(OUT_TRACE, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=[
            "MPLAN", "MEFFDATE", "MINTRATE", "MINTRATE1", "PFSA_PLAN", "PLAN_NAME"])
        w.writeheader()
        for r in records:
            w.writerow({k: r[k] for k in w.fieldnames})


def write_append_csv(records):
    """Schema-exact CSV for appending into QUIKAINT (4 DBF fields only).

    MEFFDATE is emitted as YYYYMMDD to match the DBF date storage; rates carry
    4 decimals matching N(7,4).
    """
    with open(OUT_APPEND, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["MPLAN", "MEFFDATE", "MINTRATE", "MINTRATE1"])
        for r in records:
            w.writerow([
                r["MPLAN"],
                r["MEFFDATE"],
                f"{r['MINTRATE']:.4f}",
                f"{r['MINTRATE1']:.4f}",
            ])


def main():
    entries = build_entries()
    records = build_records(entries)
    write_dbf(records)
    write_trace(records)
    write_append_csv(records)

    plans = sorted({r["MPLAN"] for r in records})
    print(f"Rate sets used: {len(entries)}")
    print(f"QLAdmin plan codes emitted: {len(plans)}")
    print(f"Records written: {len(records)} -> {OUT_DBF}")
    print(f"Append CSV: {OUT_APPEND}")
    print(f"Trace: {OUT_TRACE}")
    print("Plans:", ", ".join(plans))
    return 0


if __name__ == "__main__":
    sys.exit(main())
